import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import re
import os

# ==========================================
# 1. SYSTEM CONFIGURATION
# ==========================================
FEASIBILITY_FILE = "Door Feasbility 92187158_new.xlsx"
TOOLING_TEMPLATE = "tooling list temp.xlsx"
TOOLING_OUTPUT = "Fixture_Master_List_GENERATED.xlsx"

IMAGE_DIR = "./fixture_images/"
os.makedirs(IMAGE_DIR, exist_ok=True)

# ==========================================
# 2. IMAGE EXTRACTION PIPELINE (OPENPYXL)
# ==========================================
def extract_images_from_feasibility(filepath):
    print(f"📸 Scanning {filepath} for embedded fixture photos...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    
    # Find which column is "Fixture No" to use as our mapping key
    fix_no_col_idx = None
    for col in range(1, sheet.max_column + 1):
        cell_val = str(sheet.cell(row=4, column=col).value).upper() # Row 4 usually has headers
        if 'FIXTURE NO' in cell_val:
            fix_no_col_idx = col
            break
            
    if not fix_no_col_idx:
        print("   ⚠️ Could not locate Fixture No column for image mapping.")
        return {}

    fixture_image_map = {}
    
    for idx, image in enumerate(sheet._images):
        try:
            base_row_idx = image.anchor._from.row + 1 
            
            # 🛡️ THE FIX: Fuzzy Search for Floating Images
            # We scan the current row, then +/- 1 row, then +/- 2 rows to find the Fixture No.
            fix_no = None
            search_row_used = base_row_idx
            
            for offset in [0, 1, -1, 2, -2]:
                search_row = base_row_idx + offset
                if search_row > 0:
                    val = str(sheet.cell(row=search_row, column=fix_no_col_idx).value).replace('\n', ' ').strip()
                    if val and val.lower() not in ['none', 'nan', '']:
                        fix_no = val
                        search_row_used = search_row
                        break # Successfully locked onto a Fixture Number!
            
            if fix_no:
                img_data = None
                if hasattr(image, 'ref'):
                    img_data = image.ref.getvalue() if hasattr(image.ref, 'getvalue') else image.ref.read()
                elif hasattr(image, '_data'):
                    img_data = image._data() if callable(image._data) else image._data
                
                if img_data:
                    img_path = os.path.join(IMAGE_DIR, f"fixture_{idx}.png")
                    with open(img_path, "wb") as f:
                        f.write(img_data)
                    fixture_image_map[fix_no] = img_path
                    print(f"   ✅ Extracted image for Fixture: {fix_no} (Found near row {search_row_used})")
            else:
                print(f"   ⚠️ Found an image near row {base_row_idx}, but couldn't find a matching Fixture Number near it.")
                
        except Exception as e:
            print(f"   ⚠️ Failed to extract an image: {e}")
            
    return fixture_image_map

# ==========================================
# 3. HARVEST & INJECT PIPELINE
# ==========================================
def generate_tooling_list():
    print(f"\n🔍 Extracting Tabular Data from: {FEASIBILITY_FILE} ...")
    
    # --- 1. Extract Images First ---
    fixture_image_map = extract_images_from_feasibility(FEASIBILITY_FILE)
    
    # --- 2. Read Document Text ---
    if FEASIBILITY_FILE.endswith('.xlsx'):
        df = pd.read_excel(FEASIBILITY_FILE, sheet_name=0, skiprows=3)
    else:
        df = pd.read_csv(FEASIBILITY_FILE, skiprows=3)
        
    df.columns = df.columns.astype(str).str.strip().str.replace('\n', '')
    
    try:
        op_col = [c for c in df.columns if 'OPERATION' in str(c).upper()][0]
        fix_no_col = [c for c in df.columns if 'FIXTURE NO' in str(c).upper()][0]
        fix_name_col = [c for c in df.columns if 'FIXTURE NAME' in str(c).upper()][0]
        part_name_col = [c for c in df.columns if 'PART NAME' == str(c).upper().strip()][0]
        weld_src_col = [c for c in df.columns if 'WELDING SOURCE' in str(c).upper()][0]
        make_col = [c for c in df.columns if 'MAKE' in str(c).upper()][0]
    except IndexError as e:
        print("❌ CRITICAL: Could not find required headers in Feasibility doc!")
        return
    
    df[op_col] = df[op_col].ffill()
    df_fixtures = df.dropna(subset=[fix_no_col])
    
    match = re.search(r'\d{8,9}', FEASIBILITY_FILE)
    model_no = match.group(0) if match else "UNKNOWN"
    
    main_part_name = "UNKNOWN"
    if not df_fixtures.empty:
        main_part_name = str(df_fixtures.iloc[-1].get(part_name_col, '')).strip()
    
    print(f"✅ Extracted Header Data - Model No: {model_no} | Line: {main_part_name}")
    
    tooling_data = []
    for idx, row in df_fixtures.iterrows():
        op_no = str(row.get(op_col, '')).strip()
        fix_no = str(row.get(fix_no_col, '')).replace('\n', ' ').strip()
        fix_name = str(row.get(fix_name_col, '')).replace('\n', ' ').strip()
        p_name = str(row.get(part_name_col, '')).replace('\n', ' ').strip()
        w_src = str(row.get(weld_src_col, '')).strip()
        make_val = str(row.get(make_col, '')).strip()
        
        if fix_no.lower() == 'nan' or fix_no == '':
            continue
            
        tooling_data.append({
            "part_name": p_name if p_name.lower() != 'nan' else "",
            "fix_no": fix_no,
            "op_no": op_no,
            "cell": "1", 
            "fix_name": fix_name if fix_name.lower() != 'nan' else "",
            "w_src": w_src if w_src.lower() != 'nan' else "Manual",
            "make": make_val if make_val.lower() != 'nan' else "",
            "qty": "1" 
        })
        
    print(f"✅ Found {len(tooling_data)} fixtures to inject.")
    
    # ==========================================
    # 4. EXCEL INJECTION 
    # ==========================================
    print("\n⚙️ Booting Excel Injector for Tooling List...")
    wb = openpyxl.load_workbook(TOOLING_TEMPLATE)
    sheet = wb.active
    
    for row in range(1, 15):
        cell_val = str(sheet[f'A{row}'].value).strip()
        if "Model No :" in cell_val or "Model No:" in cell_val:
            sheet[f'A{row}'] = f"Model No :{model_no}"
        elif "LINE :" in cell_val or "LINE:" in cell_val:
            sheet[f'A{row}'] = f"LINE : {main_part_name}"
            
    start_row = 8 
    for row in range(1, 20):
        if str(sheet[f'A{row}'].value).strip() == "Sr.No.":
            start_row = row + 1
            break
            
    for idx, item in enumerate(tooling_data):
        curr_row = start_row + idx
        sheet[f'A{curr_row}'] = idx + 1
        sheet[f'B{curr_row}'] = item['part_name']
        sheet[f'C{curr_row}'] = item['fix_no']
        sheet[f'D{curr_row}'] = item['op_no']
        sheet[f'E{curr_row}'] = item['cell']
        sheet[f'F{curr_row}'] = item['fix_name']
        sheet[f'H{curr_row}'] = item['w_src']
        sheet[f'K{curr_row}'] = item['make']
        sheet[f'L{curr_row}'] = item['qty']
        
        # --- ORIGINAL METALMAN SIZE INJECTION LOGIC ---
        fix_key = item['fix_no']
        if fix_key in fixture_image_map:
            img_path = fixture_image_map[fix_key]
            if os.path.exists(img_path):
                img = ExcelImage(img_path)
                
                # Exact original average image sizing (~470x420 to 545)
                img.width, img.height = 470, 450 
                
                # Exact original Row Height from "Fixture Master List-92187158-C.xlsx"
                sheet.row_dimensions[curr_row].height = 402.0
                
                # Exact original Column Width for "Photo"
                sheet.column_dimensions['G'].width = 110.55
                
                sheet.add_image(img, f'G{curr_row}')
                print(f"   📸 Injected HUGE Original-sized Photo into row {curr_row} for {fix_key}")
        else:
            sheet.row_dimensions[curr_row].height = 25 
        
    wb.save(TOOLING_OUTPUT)
    print(f"\n🎉 SUCCESS: Tooling list with ORIGINAL METALMAN image sizing saved to -> {TOOLING_OUTPUT}")

if __name__ == "__main__":
    generate_tooling_list()
