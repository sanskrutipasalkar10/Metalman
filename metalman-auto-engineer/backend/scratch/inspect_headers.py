import openpyxl
import os

wb_path = r"c:\Users\it\Desktop\Projects\metalman_app\metalman-auto-engineer\backend\assets\CONTROL_PLAN_TEMPLATE.xlsx"
if os.path.exists(wb_path):
    wb = openpyxl.load_workbook(wb_path)
    ws = wb["Q-Sheetmetal Parts"]
    for r in range(15, 20):
        print(f"Row {r}:")
        for col in range(1, 16):
            val = ws.cell(r, col).value
            if val is not None:
                print(f"  Col {openpyxl.utils.get_column_letter(col)}: {val}")
else:
    print("File not found")
