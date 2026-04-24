import openpyxl
import sys

# Ensure UTF-8 output for console
if sys.stdout.encoding != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'assets\test_inputs\PFD_Completed_TEST_v4.xlsx')
sheet = wb['SUB_ASSY']

# Check Part Description Label and Value
print(f"A5 (Label): '{sheet['A5'].value}'")
# Search row 5 for the part description value
for c in range(2, 10):
    val = sheet.cell(row=5, column=c).value
    if val and 'DOOR' in str(val).upper():
        print(f"Column {c} (Value): '{val}'")

# Check Symbols
print("\nChecking symbols in row 10 (SUB_ASSY):")
for c in range(2, 6):
    cell = sheet.cell(row=10, column=c)
    color = cell.font.color.rgb if cell.font.color else "None"
    print(f"Col {c}: Value='{cell.value}', Font Color='{color}', Bold={cell.font.bold}")
