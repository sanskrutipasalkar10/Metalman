import sys
import os
import glob
import subprocess
import pandas as pd
import re
import cadquery as cq
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import cv2
import numpy as np
from openpyxl.styles import PatternFill, Font

# Standard process symbol color (Soft Green) - Increased size for visibility
GREEN_FONT = Font(color='00B050', bold=True, size=28)

# ==========================================
# 1. SYSTEM CONFIGURATION
# ==========================================
def find_freecad_cmd():
    print("Locating FreeCAD Command Line Engine...")
    search_paths = glob.glob(r"C:\Program Files\FreeCAD*\bin\FreeCADCmd.exe")
    if not search_paths:
        print("WARNING: Could not find FreeCADCmd.exe in C:\\Program Files\\")
        return None
    print(f"Found Isolated FreeCAD Engine: {search_paths[0]}")
    return search_paths[0]

FREECAD_CMD = find_freecad_cmd()
WORKER_SCRIPT = "slicer_worker.py"

# ==========================================
# UTILS
# ==========================================
def safe_write_data(sheet, cell_coord, value, font=None):
    """Writes value to a cell, ensuring it doesn't break merged cells by writing to top-left."""
    cell = sheet[cell_coord]
    target_cell = cell
    if type(cell).__name__ == 'MergedCell':
        for merged_range in sheet.merged_cells.ranges:
            if cell_coord in merged_range:
                target_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    
    target_cell.value = value
    if font:
        target_cell.font = font

import datetime

def extract_level_0_headers(feasibility_xlsx):
    """Utility to grab the top-level assembly info for Headers."""
    try:
        df = pd.read_excel(feasibility_xlsx, skiprows=3, engine='openpyxl')
        df.columns = df.columns.astype(str).str.strip().str.replace('\n', '')
        
        part_number = ""
        part_desc = ""
        rev = ""
        
        for index, row in df.iterrows():
            bom_level = str(row.get('BOM Level', '')).strip()
            if bom_level == '0' or bom_level == '0.0':
                p = str(row.get('Part No.', '')).strip()
                if p.endswith('.0'): p = p[:-2]
                part_number = p
                part_desc = str(row.get('Part Description', '')).strip()
                rev = str(row.get('Revision no', '')).strip()
                break
                
        return part_number, part_desc, rev
    except Exception:
        return "", "", ""

def inject_sheet_headers(sheet, part_num, part_desc, rev, max_row=10, max_col=20):
    """Scan the top rows of a sheet for header label cells and fill them."""
    today = datetime.date.today().strftime("%d-%m-%Y")
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            raw = sheet.cell(row=r, column=c).value
            if raw is None:
                continue
            label = str(raw).strip().upper()
            col_letter = openpyxl.utils.get_column_letter(c)
            # Find the next non-empty cell to the right for value
            def write_next(lbl, val):
                # Find the first available cell to the right that isn't part of the label's merge range
                for sc in range(c + 1, max_col + 1):
                    # Check if (r, sc) is part of the same merge range as the label cell (r, c)
                    is_in_same_merge = False
                    coord_label = f"{openpyxl.utils.get_column_letter(c)}{r}"
                    coord_target = f"{openpyxl.utils.get_column_letter(sc)}{r}"
                    for merged_range in sheet.merged_cells.ranges:
                        if coord_label in merged_range and coord_target in merged_range:
                            is_in_same_merge = True
                            break
                    if is_in_same_merge:
                        continue

                    nc = sheet.cell(row=r, column=sc)
                    if nc.value is None or str(nc.value).strip() == '':
                        safe_write_data(sheet, f"{openpyxl.utils.get_column_letter(sc)}{r}", val)
                        return
                    elif str(nc.value).strip() == str(val).strip():
                        return
            if 'PART DESCRIPTION' in label or label == 'PART DESCRIPTION':
                write_next(label, part_desc)
            elif 'PART NAME' in label:
                write_next(label, part_desc)
            elif 'PART NO' in label or 'PART NUMBER' in label:
                write_next(label, part_num)
            elif 'REVISION' in label and 'NUMBER' not in label:
                write_next(label, rev)
            elif 'DATE' in label and 'UPDATE' not in label:
                write_next(label, today)

# ==========================================
# PFD SHEET 1: SUB-ASSEMBLY INDEX LOGIC
# ==========================================
def resolve_phantom_parts(current_op, parts_dictionary, phantoms_dictionary, visited=None):
    if visited is None: visited = set()
    if current_op in visited: return [] # Avoid circular refs
    visited.add(current_op)
    
    resolved_parts = set()
    parts_here = parts_dictionary.get(current_op, [])
    resolved_parts.update(parts_here)
    
    related_phantoms = phantoms_dictionary.get(current_op, [])
    if related_phantoms:
        for phantom_ref in related_phantoms:
            phantom_ref_parts = resolve_phantom_parts(phantom_ref, parts_dictionary, phantoms_dictionary, visited.copy())
            resolved_parts.update(phantom_ref_parts)
            
    return list(resolved_parts)


def extract_feasibility_data(filepath):
    print(f"   [FEASIBILITY] Reading file: {filepath}...", flush=True)
    if filepath.endswith('.xlsx'):
        df = pd.read_excel(filepath, sheet_name=0, skiprows=3)
    else:
        df = pd.read_csv(filepath, skiprows=3)
    print(f"   [FEASIBILITY] File read complete. Processing columns...", flush=True)
        
    df.columns = df.columns.astype(str).str.strip().str.replace('\n', '')
    
    try:
        op_col = [c for c in df.columns if 'OPERATION' in str(c).upper()][0]
        part_no_col = [c for c in df.columns if 'PART NO' in str(c).upper()][0]
        fixture_col = [c for c in df.columns if 'FIXTURE NAME' in str(c).upper()][0]
        phantom_col = [c for c in df.columns if 'SUB ASSEMBLIES USED' in str(c).upper()][0]
    except IndexError:
        print(f"   [FEASIBILITY] ERROR: Could not find required columns in feasibility sheet.")
        return {}, []
    
    df[op_col] = df[op_col].ffill()
    
    parts_dictionary = {}
    phantoms_dictionary = {}
    
    # 1. First Pass: Map raw parts per operation
    raw_parts_df = df.dropna(subset=[part_no_col]).copy()
    raw_parts_df.loc[:, part_no_col] = raw_parts_df[part_no_col].astype(str).str.replace('.0', '', regex=False).str.strip()
    raw_parts_df = raw_parts_df[raw_parts_df[part_no_col].str.lower() != 'nan']
    raw_parts_df = raw_parts_df[raw_parts_df[part_no_col] != '']
    
    parts_dictionary = raw_parts_df.groupby(op_col)[part_no_col].apply(list).to_dict()
    
    print(f"   [FEASIBILITY] Extracted {len(parts_dictionary)} operations. Resolving phantoms...", flush=True)
    
    # 2. Second Pass: Map phantom references (Sub assemblies used)
    phantom_rows = df.dropna(subset=[phantom_col])
    
    for _, row in phantom_rows.iterrows():
        op_no = str(row.get(op_col, '')).strip()
        phantoms_str = str(row.get(phantom_col, '')).strip()
        
        phantom_list = [p.strip() for p in phantoms_str.split(',') if p.strip()]
        
        if op_no in phantoms_dictionary:
            phantoms_dictionary[op_no].extend(phantom_list)
        else:
            phantoms_dictionary[op_no] = phantom_list
            
    dynamic_routing_matrix = {}
    for current_op in parts_dictionary.keys():
        resolved_list = resolve_phantom_parts(current_op, parts_dictionary, phantoms_dictionary)
        dynamic_routing_matrix[current_op] = resolved_list

    for op in phantoms_dictionary.keys():
        if op not in dynamic_routing_matrix:
             dynamic_routing_matrix[op] = resolve_phantom_parts(op, parts_dictionary, phantoms_dictionary)

    sub_assy_data = []
    op_rows = df.dropna(subset=[fixture_col])
    
    for index, row in op_rows.iterrows():
        op_no = str(row.get(op_col, '')).strip()
        fixture_name = str(row.get(fixture_col, '')).strip()
        
        if not op_no or op_no.lower() == 'nan' or not fixture_name or fixture_name.lower() == 'nan':
            continue
            
        part_no_match = re.search(r'\b(\d{8,9}(?:-\d+)?)\b', fixture_name)
        extracted_part_no = part_no_match.group(1) if part_no_match else ""
        
        rev_match = re.search(r'REV-([A-Z])', fixture_name.upper())
        rev = rev_match.group(1) if rev_match else ""

        sub_assy_data.append({
            "operation": op_no,
            "part_name": fixture_name,
            "part_no": extracted_part_no,
            "rev": rev
        })

    return dynamic_routing_matrix, sub_assy_data

def isolate_part_and_remove_text(image_path, output_filename, padding=10):
    img = cv2.imread(image_path)
    if img is None: return False
    
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY_INV)
    
    kernel = np.ones((7, 7), np.uint8) 
    dilated = cv2.dilate(thresh, kernel, iterations=1)
    
    contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if not contours:
        return False
    
    best_cnt = max(contours, key=cv2.contourArea)
    x, y, w, h = cv2.boundingRect(best_cnt)
    
    x_start = max(0, x - padding)
    y_start = max(0, y - padding)
    x_end = min(img.shape[1], x + w + padding)
    y_end = min(img.shape[0], y + h + padding)
    
    final_cropped = img[y_start:y_end, x_start:x_end]
    cv2.imwrite(output_filename, final_cropped)
    return True

def render_temporary_step(temp_step_path, stage_name, output_dir):
    try:
        shape = cq.importers.importStep(temp_step_path).val()
        
        render_options = {
            "width": 1600, "height": 1200, 
            "marginLeft": 20, "marginTop": 20,
            "showAxes": False,
            "projectionDir": (1, -1, 1), 
            "strokeWidth": 2.5,          
            "strokeColor": (0, 0, 0),    
            "hiddenColor": (200, 200, 200), 
            "showHidden": True           
        }
        
        safe_stage_name = stage_name.replace(" ", "_")
        svg_filename = os.path.join(output_dir, f"{safe_stage_name}.svg")
        png_raw_filename = os.path.join(output_dir, f"{safe_stage_name}_raw.png")
        png_clean_filename = os.path.join(output_dir, f"{safe_stage_name}.png")
        
        cq.exporters.export(shape, svg_filename, opt=render_options)
        
        try:
            import fitz  
            svg_doc = fitz.open(svg_filename)
            # Force a white background across the SVG by dropping the Alpha channel.
            # Without this, transparent SVGs render as black in OpenCV, destroying the threshold.
            pix = svg_doc.load_page(0).get_pixmap(dpi=200, alpha=False)
            pix.save(png_raw_filename)
            svg_doc.close()
            
            isolation_success = isolate_part_and_remove_text(png_raw_filename, png_clean_filename)
            
            if isolation_success:
                if os.path.exists(png_raw_filename):
                    os.remove(png_raw_filename) 
            else:
                os.rename(png_raw_filename, png_clean_filename)
            
            print(f"      - PNG rendered successfully: {png_clean_filename}", flush=True)
                
        except Exception as e:
            print(f"   [RENDER ERROR] PNG Conversion failed: {e}. Saved as SVG instead.", flush=True)

    except Exception as e:
        print(f"Render failed for {stage_name}: {e}")

def generate_cad_images(routing_matrix, master_step_file, output_dir):
    if not FREECAD_CMD:
        print("   [CAD] ERROR: FreeCAD command not found, skipping rendering.")
        return

    abs_master = os.path.abspath(master_step_file)
    abs_output_dir = os.path.abspath(output_dir)
    
    # 1. Filter valid stages
    valid_routing = {k: v for k, v in routing_matrix.items() if v}
    if not valid_routing:
        print("   [CAD] No valid stages to render.")
        return

    print(f"   [CAD] Batch extracting {len(valid_routing)} stages in a single FreeCAD session...", flush=True)
    
    local_worker_script = os.path.join(abs_output_dir, WORKER_SCRIPT)
    escaped_master = abs_master.replace("\\", "\\\\")
    
    # Create map of target files for the worker
    stage_file_map = {}
    for stage_name in valid_routing.keys():
        safe_stage_name = stage_name.replace(" ", "_")
        abs_temp_step = os.path.join(abs_output_dir, f"temp_{safe_stage_name}.stp")
        stage_file_map[stage_name] = abs_temp_step.replace("\\", "\\\\")

    worker_code = f"""
import FreeCAD
import Import
import sys
import os

param = FreeCAD.ParamGet("User parameter:BaseApp/Preferences/Mod/Import")
param.SetBool("ExpandCompound", True)
param.SetBool("UseOCAF", True) 

master_file = r"{escaped_master}"
routing_matrix = {valid_routing}
stage_file_map = {stage_file_map}

print(f"Loading master file: {{master_file}}")
doc = FreeCAD.newDocument("MasterDoc")
Import.insert(master_file, doc.Name)

all_objs = doc.Objects
obj_labels = []
for obj in all_objs:
    if hasattr(obj, 'Label'):
        obj_labels.append((obj, str(obj.Label)))

for stage_name, target_parts in routing_matrix.items():
    output_file = stage_file_map.get(stage_name)
    if not output_file: continue
    
    print(f"Extracting stage: {{stage_name}}")
    objects_to_export = []
    for obj, label in obj_labels:
        for part in target_parts:
            if str(part).strip() in label:
                objects_to_export.append(obj)
                break
    
    if objects_to_export:
        Import.export(objects_to_export, output_file)
        print(f"Exported {{stage_name}} to {{output_file}}")
    else:
        print(f"No objects found for {{stage_name}}")

sys.exit(0)
"""
    with open(local_worker_script, "w", encoding="utf-8") as f:
        f.write(worker_code)
    
    # Run the batch extraction
    result = subprocess.run([FREECAD_CMD, local_worker_script], capture_output=True, text=True)
    
    if result.returncode != 0:
        print(f"   [CAD] Batch FreeCAD extraction failed. Code: {result.returncode}")
        print(f"   [DEBUG] {result.stderr.strip()[-500:]}")
    else:
        print(f"   [CAD] Batch extraction successful. Now rendering PNGs...")
        
    # 2. Render PNGs using CadQuery for each extracted STEP
    for i, (stage_name, output_file) in enumerate(stage_file_map.items()):
        abs_temp_step = os.path.normpath(output_file)
        if os.path.exists(abs_temp_step):
            print(f"   [CAD] [{i+1}/{len(stage_file_map)}] Rendering PNG for {stage_name}...")
            render_temporary_step(abs_temp_step, stage_name, abs_output_dir)
            os.remove(abs_temp_step)
        else:
            print(f"   [CAD] Missing step file for {stage_name}: {abs_temp_step}")
            
    if os.path.exists(local_worker_script):
        os.remove(local_worker_script)
    
    print(f"   [CAD] All images generated for PFD Sheet 1.", flush=True)




def write_sub_assy_index(assy_data, pfd_template, pfd_output, output_dir):
    if not assy_data: return

    wb = openpyxl.load_workbook(pfd_template)
    if "Sub Assmbly Index" not in wb.sheetnames: return
        
    sheet = wb["Sub Assmbly Index"]
    start_row = 3
    
    for idx, item in enumerate(assy_data):
        current_row = start_row + idx
        
        sheet[f'A{current_row}'] = idx + 1
        sheet[f'B{current_row}'] = item["operation"]
        sheet[f'C{current_row}'] = item["part_name"]
        sheet[f'D{current_row}'] = item["part_no"]
        if item["rev"]: sheet[f'E{current_row}'] = item["rev"]

        safe_stage_name = str(item["operation"]).replace(" ", "_")
        image_path = os.path.join(output_dir, f"{safe_stage_name}.png")
        
        if os.path.exists(image_path):
            img = ExcelImage(image_path)
            sheet.row_dimensions[current_row].height = 142
            sheet.column_dimensions['F'].width = 35
            img.width, img.height = 190, 140 
            sheet.add_image(img, f'F{current_row}')
        else:
            sheet.row_dimensions[current_row].height = 25 

    wb.save(pfd_output)

def render_sub_assy_index(feasibility_file, master_step_file, pfd_template, pfd_output, output_dir):
    """SHEET 1 ORCHESTRATOR"""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        
    routing_matrix, sub_assy_data = extract_feasibility_data(feasibility_file)
    generate_cad_images(routing_matrix, master_step_file, output_dir)
    write_sub_assy_index(sub_assy_data, pfd_template, pfd_output, output_dir)
    return {"status": "success", "sub_assy_index_path": pfd_output}

# ==========================================
# PFD SHEET 2: SUB_ASSY ROUTER & INJECTOR
# ==========================================
def extract_fixture_mapping_from_feasibility(feasibility_xlsx):
    try:
        df = pd.read_excel(feasibility_xlsx, skiprows=3, engine='openpyxl')
        df.columns = df.columns.astype(str).str.strip().str.replace('\n', '')
        
        try:
            op_col = [c for c in df.columns if 'OPERATION' in str(c).upper()][0]
            part_col = [c for c in df.columns if 'PART NO' in str(c).upper()][0]
            desc_col = [c for c in df.columns if 'PART DESCRIPTION' in str(c).upper()][0] if any('PART DESCRIPTION' in str(c).upper() for c in df.columns) else None
            sub_assy_col = [c for c in df.columns if 'SUB ASSEMBLIES USED' in str(c).upper()][0]
        except IndexError:
            return {}

        df[op_col] = df[op_col].ffill()
        grouped = df.groupby(op_col)
        
        # 🚀 CHRONOLOGICAL STATE MACHINE: Tracks the timeline based on position in file
        all_ops_ordered = []
        for op in df[op_col].dropna():
            clean = str(op).strip().upper().replace(" ", "")
            if clean and clean != 'NAN' and clean not in all_ops_ordered:
                all_ops_ordered.append(clean)
                
        grouped = df.groupby(op_col)
        operation_map = {}
        
        for raw_op, group in grouped:
            clean_op = str(raw_op).strip().upper().replace(" ", "")
            if clean_op == 'NAN' or not clean_op: continue
            
            # Identify immediate true processing predecessor
            try:
                op_idx = all_ops_ordered.index(clean_op)
                predecessor = all_ops_ordered[op_idx - 1] if op_idx > 0 else None
            except ValueError:
                predecessor = None

            parts_list = []
            for _, row in group.iterrows():
                p = str(row.get(part_col, '')).strip().replace('.0', '')
                if p and p.lower() != 'nan': parts_list.append(p)
            
            # 🚀 SEQUENTIAL INHERITANCE: Automatically binds predecessor output 
            if predecessor and not clean_op.startswith("OP-10"):
                display_pred = predecessor
                # Formatting: adds uniform spaces for downstream readability (e.g. OP-10 B)
                if display_pred.startswith("OP-") and len(display_pred) > 5 and display_pred[5].isalpha():
                    display_pred = display_pred[:5] + " " + display_pred[5:]
                
                sub_list = [f"{display_pred}-1Nos"]
            else:
                # Fallback path for initial operations (OP-10 level setup layers)
                sub_list = []
                for _, row in group.iterrows():
                    s = str(row.get(sub_assy_col, '')).strip()
                    if s and s.lower() != 'nan':
                        for split_s in s.split(','):
                            if split_s.strip(): sub_list.append(split_s.strip())

            # Format outputs
            parts_block = "Part No. :\n" + "\n".join(set(parts_list)) if parts_list else ""
            
            if sub_list:
                formatted_subs = []
                for s in set(sub_list):
                    if "-1NOS" not in s.upper() and any(kw in s.upper() for kw in ["OP-", "ASSY"]):
                        formatted_subs.append(f"{s}-1Nos")
                    else:
                        formatted_subs.append(s)
                sub_block = "Sub Assemblies :\n" + "\n".join(formatted_subs)
            else:
                sub_block = ""

            # Standardized layout routing conditions
            if clean_op == "OP-10A":
                payload = parts_block
            elif clean_op == "OP-10B" or clean_op == "OP-20":
                payload = (sub_block + "\n\n" + parts_block).strip()
            elif clean_op == "OP-30":
                # Forced layout for OP-30
                payload = sub_block if sub_block else parts_block
            elif clean_op == "OP-50":
                payload = ""
            else:
                payload = (sub_block + "\n\n" + parts_block).strip()
                
            operation_map[clean_op] = payload

        return operation_map
    except Exception as e:
        print(f"extract_fixture_mapping error: {e}")
        return {}

def _get_process_symbol_columns(sheet, max_row=15, max_col=25):
    """Scan sheet headers to find column indexes for OPERATION/MOVEMENT/INSPECTION/STORE."""
    symbol_cols = {"operation": None, "movement": None, "inspection": None, "store": None}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = sheet.cell(row=r, column=c).value
            if val is None: continue
            t = str(val).strip().upper()
            if t == "OPERATION" and symbol_cols["operation"] is None:
                symbol_cols["operation"] = c
            elif t == "MOVEMENT" and symbol_cols["movement"] is None:
                symbol_cols["movement"] = c
            elif t == "INSPECTION" and symbol_cols["inspection"] is None:
                symbol_cols["inspection"] = c
            elif t == "STORE" and symbol_cols["store"] is None:
                symbol_cols["store"] = c
        if all(v is not None for v in symbol_cols.values()):
            break
    return symbol_cols

def _classify_row_process(desc_text):
    """Returns which process type a row is based on its description."""
    d = desc_text.upper().replace(' ', '').replace('\n', '')
    if any(k in d for k in ['MOVE', 'TRANSPORT', 'TROLLEY', 'TRANSFER', 'FORKLIFT', 'CARRY']):
        return 'movement'
    if any(k in d for k in ['INSPECT', 'CHECK', 'QUALITY', 'QC', 'VERIFY', 'DIMENSION', 'MEASURE']):
        return 'inspection'
    if any(k in d for k in ['STORE', 'STORAGE', 'RACK', 'RETRIEVE', 'ISSUE', 'BUFFER']):
        return 'store'
    return 'operation'  # default

def render_sub_assy_sheet(feasibility_file, pfd_template, pfd_output):
    """SHEET 2 ORCHESTRATOR"""
    TARGET_SHEET = "SUB_ASSY"
    PARTS_TARGET_COL = 'H'
    
    # Standard process symbols (Unicode)
    SYMBOLS = {
        'operation': u'\u25CF',   # ● filled circle
        'movement':  u'\u25BA',   # ► filled right arrow
        'inspection': u'\u25A0',  # ■ filled square
        'store':     u'\u25BC',   # ▼ filled triangle
    }

    part_num, part_desc, rev = extract_level_0_headers(feasibility_file)
    operation_map = extract_fixture_mapping_from_feasibility(feasibility_file)
    if not operation_map:
        return {"status": "error", "message": "Failed to map operations."}

    wb = openpyxl.load_workbook(pfd_template)
    
    if TARGET_SHEET not in wb.sheetnames:
        return {"status": "error", "message": f"Sheet {TARGET_SHEET} not found."}
        
    sheet = wb[TARGET_SHEET]
    
    # Inject consistent headers
    inject_sheet_headers(sheet, part_num, part_desc, rev)

    # Find the OPERATION/MOVEMENT/INSPECTION/STORE symbol columns
    sym_cols = _get_process_symbol_columns(sheet)

    current_major_op = None

    for row_num in range(8, 250): 
        val_a = sheet[f'A{row_num}'].value
        if val_a is None: continue
        
        cell_text = str(val_a).strip().upper()
        
        if cell_text.startswith("OP-"):
            clean_target_op = cell_text.replace(" ", "")
            current_major_op = clean_target_op if clean_target_op in operation_map else None
            continue 
            
        if "CFT TEAM MEMBERS" in cell_text or "REVISION HISTORY" in cell_text or "PREPARED BY" in cell_text:
            break
        # Also check column C just in case
        val_c = str(sheet[f'C{row_num}'].value).strip().upper() if sheet[f'C{row_num}'].value else ""
        if "CFT TEAM MEMBERS" in val_c or "REVISION HISTORY" in val_c or "PREPARED BY" in val_c:
            break

        # Inject process symbols based on F-column description
        desc_val = sheet[f'F{row_num}'].value
        desc_text = str(desc_val) if desc_val else ""
        if desc_text.strip():
            proc_type = _classify_row_process(desc_text)
            sym_col = sym_cols.get(proc_type)
            if sym_col:
                col_letter = openpyxl.utils.get_column_letter(sym_col)
                cell_ref = f"{col_letter}{row_num}"
                safe_write_data(sheet, cell_ref, SYMBOLS[proc_type], font=GREEN_FONT)

        if current_major_op:
            # OP-30: write sub-assembly payload only on first matching step row
            if current_major_op == "OP-30":
                if cell_text.replace(".", "").isdigit():
                    step_num = int(float(cell_text))
                    # Only write on first data step (step 10)
                    if step_num == 10:
                        payload = operation_map.get("OP-30", "")
                        if payload:
                            safe_write_data(sheet, f'{PARTS_TARGET_COL}{row_num}', payload)
                        current_major_op = None
            else:
                if cell_text in ("10.0", "10"):
                    if "CHILD" in desc_text.upper() or "PART" in desc_text.upper() or "COLLECTION" in desc_text.upper() or "INSPECTION" in desc_text.upper():
                        payload = operation_map[current_major_op]
                        safe_write_data(sheet, f'{PARTS_TARGET_COL}{row_num}', payload)
                        current_major_op = None 

    wb.save(pfd_output)
    return {"status": "success", "sub_assy_path": pfd_output}

# ==========================================
# PFD SHEET 3: SHEETMETAL INJECTOR
# ==========================================
def extract_and_segregate_sheetmetal(feasibility_file):
    try:
        df = pd.read_excel(feasibility_file, skiprows=3, engine='openpyxl')
        df.columns = df.columns.astype(str).str.strip().str.replace('\n', '')
        
        # Find columns with flexible matching (handle MANUFACTCURING typo, etc)
        part_col_matches = [c for c in df.columns if 'PART NO' in c.upper()]
        process_col_matches = [c for c in df.columns if 'MANUFACTUR' in c.upper() or 'PROCESS DETAIL' in c.upper()]
        qty_col_matches = [c for c in df.columns if 'QTY' in c.upper()]
        remarks_col_matches = [c for c in df.columns if 'COMMODITY' in c.upper() or 'REMARK' in c.upper()]
        
        if not all([part_col_matches, process_col_matches, qty_col_matches, remarks_col_matches]):
            print(f"Sheetmetal: Missing columns. Found process_cols={process_col_matches}")
            return None
        
        part_col = part_col_matches[0]
        process_col = process_col_matches[0]
        qty_col = qty_col_matches[0]
        remarks_col = remarks_col_matches[0]

        routed_parts = {
            "all": [], "laser": [], "oxy": [], "deburr": [], 
            "straighten": [], "bend": [], "drill": [], "tap": [], "chamfer": []
        }
        
        for index, row in df.iterrows():
            raw_part = str(row.get(part_col, ''))
            if not raw_part or raw_part.lower() == 'nan': continue
            
            part_number = re.sub(r'\s+', '', raw_part)
            if part_number.endswith('.0'): part_number = part_number[:-2]
            
            remarks = str(row.get(remarks_col, '')).strip().upper()
            
            if 'SHEETMETAL' in remarks or 'SHEET METAL' in remarks:
                raw_qty = str(row.get(qty_col, '1'))
                qty = re.sub(r'\s+', '', raw_qty)
                if qty.endswith('.0'): qty = qty[:-2]
                
                formatted_part = f"{part_number}-{qty}Nos"
                process = str(row.get(process_col, '')).upper()
                
                routed_parts["all"].append(formatted_part)
                
                # LC / LASER
                if re.search(r'(LC|LASER)', process): routed_parts["laser"].append(formatted_part)
                # GS = Gas Shearing = Oxy/Plasma equivalent; also OXY, PLASMA
                if re.search(r'(OXY|PLASMA|GS)', process): routed_parts["oxy"].append(formatted_part)
                # GRIND / DEBURR / GRINDING
                if re.search(r'(DEBURR|GRIND)', process): routed_parts["deburr"].append(formatted_part)
                # STRAIGHTEN or STR abbreviation
                if re.search(r'(STRAIGHTEN|\bSTR\b)', process): routed_parts["straighten"].append(formatted_part)
                # BEND / FORM / BENDING
                if re.search(r'(BEND|FORM)', process): routed_parts["bend"].append(formatted_part)
                # DRILL and also DRLL (common site typo in the data)
                if re.search(r'(DRILL|DRLL)', process): routed_parts["drill"].append(formatted_part)
                # TAP / TAPPING
                if re.search(r'(TAP)', process): routed_parts["tap"].append(formatted_part)
                # CHAMFER / CHA
                if re.search(r'(CHAMFER|\bCHA\b)', process): routed_parts["chamfer"].append(formatted_part)

        return routed_parts
    except Exception as e:
        return None

def render_sheetmetal_sheet(feasibility_file, pfd_template, pfd_output):
    """SHEET 3 ORCHESTRATOR"""
    SM_SHEET_NAME = "Q-SHEETMETAL_PARTS"
    TARGET_LIST_COL = 'H' 
    
    # Dynamically extract headers
    part_num, part_desc, rev = extract_level_0_headers(feasibility_file)
    HEADER_PART_DESC = part_desc if part_desc else "ASSEMBLY"
    HEADER_PART_NUM = part_num if part_num else "UNKNOWN"
    HEADER_REV_NUM = f'Revision Number :- "{rev}"' if rev else 'Revision Number :- "0"'

    routed_parts = extract_and_segregate_sheetmetal(feasibility_file)
    if not routed_parts:
        return {"status": "error", "message": "No sheetmetal parts found."}

    raw_text = "Sheetmetal Parts :\n" + "\n".join(routed_parts["all"])
    laser_text = "Laser Cutting Parts List :\n\n" + "\n".join(routed_parts["laser"])
    oxy_text = "Oxy / Plasma Cutting Parts List :\n\n" + "\n".join(routed_parts["oxy"])
    deburr_text = "Deburring / Grinding Parts List :\n\n" + "\n".join(routed_parts["deburr"])
    straighten_text = "Straightening Parts List :\n\n" + "\n".join(routed_parts["straighten"])
    bend_text = "Bending Parts List :\n" + "\n".join(routed_parts["bend"])
    drill_text = "Drilling Parts List :\n\n" + "\n".join(routed_parts["drill"])
    tap_text = "Tapping Parts List :\n\n" + "\n".join(routed_parts["tap"])
    chamfer_text = "Chamfering Parts List :\n\n" + "\n".join(routed_parts["chamfer"])
    storage_text = "Parts List for Storage :\n" + "\n".join(routed_parts["all"])

    wb = openpyxl.load_workbook(pfd_template)
    
    if SM_SHEET_NAME not in wb.sheetnames:
        return {"status": "error", "message": "Sheetmetal sheet not found"}

    sheet_obj = wb[SM_SHEET_NAME]
    
    # Inject standard headers + process symbols
    inject_sheet_headers(sheet_obj, HEADER_PART_NUM, HEADER_PART_DESC, rev)
    sym_cols = _get_process_symbol_columns(sheet_obj)
    SYMBOLS = {
        'operation': u'\u25CF', 'movement':  u'\u25BA',
        'inspection': u'\u25A0', 'store': u'\u25BC',
    }

    # Legacy specific header hunt (row 5 labels)
    for c in range(1, 20):
        cell_str = str(sheet_obj.cell(row=5, column=c).value).strip() if sheet_obj.cell(row=5, column=c).value else ''
        if cell_str == "Part Description":
            for search_col in range(c + 1, 20):
                if not sheet_obj.cell(row=5, column=search_col).value:
                    safe_write_data(sheet_obj, f"{openpyxl.utils.get_column_letter(search_col)}5", HEADER_PART_DESC)
                    break
        elif cell_str == "Part Number":
            for search_col in range(c + 1, 20):
                if not sheet_obj.cell(row=5, column=search_col).value:
                    safe_write_data(sheet_obj, f"{openpyxl.utils.get_column_letter(search_col)}5", HEADER_PART_NUM)
                    break
        elif "Revision Number" in cell_str:
            safe_write_data(sheet_obj, f"{openpyxl.utils.get_column_letter(c)}5", HEADER_REV_NUM)
            

    for row_num in range(8, 150): 
        val_a = sheet_obj[f'A{row_num}'].value
        if val_a is None: continue
        
        cell_text = str(val_a).strip().upper()
        if "CFT TEAM MEMBERS" in cell_text or "REVISION HISTORY" in cell_text or "PREPARED BY" in cell_text:
            break
        # Check column C too
        val_c = str(sheet_obj[f'C{row_num}'].value).strip().upper() if sheet_obj[f'C{row_num}'].value else ""
        if "CFT TEAM MEMBERS" in val_c or "REVISION HISTORY" in val_c or "PREPARED BY" in val_c:
            break
        
        op_num = str(val_a).strip()
        if op_num.endswith('.0'): op_num = op_num[:-2]
        if not op_num.isdigit(): continue  

        val_f = sheet_obj[f'F{row_num}'].value
        if not val_f: continue
        
        op_desc = str(val_f).upper().replace(' ', '').replace('\n', '').replace('\xa0', '')
        
        if "RAWMATERIAL" in op_desc or "INCOMING" in op_desc or "ISSUEMATERIAL" in op_desc or "SHEARING" in op_desc:
            safe_write_data(sheet_obj, f'{TARGET_LIST_COL}{row_num}', raw_text)
        elif "LASER" in op_desc:
            safe_write_data(sheet_obj, f'{TARGET_LIST_COL}{row_num}', laser_text)
        elif "DRILL" in op_desc:
            safe_write_data(sheet_obj, f'{TARGET_LIST_COL}{row_num}', drill_text)
        elif "TAP" in op_desc:
            safe_write_data(sheet_obj, f'{TARGET_LIST_COL}{row_num}', tap_text)
        elif "CHAMFER" in op_desc:
            safe_write_data(sheet_obj, f'{TARGET_LIST_COL}{row_num}', chamfer_text)
        elif "OXY" in op_desc or "PLASMA" in op_desc:
            safe_write_data(sheet_obj, f'{TARGET_LIST_COL}{row_num}', oxy_text)
        elif "DEBURR" in op_desc or "GRIND" in op_desc:
            safe_write_data(sheet_obj, f'{TARGET_LIST_COL}{row_num}', deburr_text)
        elif "STRAIGHTEN" in op_desc:
            safe_write_data(sheet_obj, f'{TARGET_LIST_COL}{row_num}', straighten_text)
        elif "BEND" in op_desc or "FORMING" in op_desc:
            safe_write_data(sheet_obj, f'{TARGET_LIST_COL}{row_num}', bend_text)
        elif "STORAGE" in op_desc:
            safe_write_data(sheet_obj, f'{TARGET_LIST_COL}{row_num}', storage_text)
        elif "MOVE" in op_desc:
            safe_write_data(sheet_obj, f'{TARGET_LIST_COL}{row_num}', "")

        # Inject process symbol for this row
        val_f = sheet_obj[f'F{row_num}'].value
        desc_text = str(val_f) if val_f else ""
        if desc_text.strip():
            proc_type = _classify_row_process(desc_text)
            sym_col = sym_cols.get(proc_type)
            if sym_col:
                col_letter = openpyxl.utils.get_column_letter(sym_col)
                cell_ref = f"{col_letter}{row_num}"
                safe_write_data(sheet_obj, cell_ref, SYMBOLS[proc_type], font=GREEN_FONT)

    wb.save(pfd_output)
    return {"status": "success", "sheetmetal_path": pfd_output}

# ==========================================
# PFD SHEET 4: BOP INJECTOR
# ==========================================
def extract_bop_parts(feasibility_xlsx):
    try:
        df = pd.read_excel(feasibility_xlsx, skiprows=3, engine='openpyxl')
        df.columns = df.columns.astype(str).str.strip().str.replace('\n', '')
        
        bom_data = []
        for index, row in df.iterrows():
            if pd.isna(row.get('Sr No')) or str(row.get('Sr No')).strip() == '': continue
                
            part_number = str(row.get('Part No.')).strip()
            if part_number.endswith('.0'): part_number = part_number[:-2]

            qty = str(row.get('Qty/Assy')).strip()
            if qty.endswith('.0'): qty = qty[:-2]
            
            bom_data.append({
                "part_no": part_number,
                "qty": qty,
                "remarks": str(row.get('Commodity', ''))
            })
            
        return bom_data
    except Exception:
        return []

def render_bop_sheet(feasibility_file, pfd_template, pfd_output):
    """SHEET 4 ORCHESTRATOR"""
    BOP_SHEET_NAME = "T - BOP & Hardwares"
    L0_DESC_CELL = 'C5'
    L0_PART_CELL = 'G5'
    L0_REV_CELL  = 'I5'
    BOP_TARGET_COL = 'H'  
    SYMBOLS = {
        'operation': u'\u25CF', 'movement':  u'\u25BA',
        'inspection': u'\u25A0', 'store': u'\u25BC',
    }

    part_num, part_desc, rev = extract_level_0_headers(feasibility_file)
    all_bom_parts = extract_bop_parts(feasibility_file)
    
    bop_keywords = ['HARDWARE', 'BOP', 'FASTENER', 'BOUGHT OUT', 'STANDARD']
    bop_list = [item for item in all_bom_parts if any(kw in item['remarks'].upper() for kw in bop_keywords)]

    op10_lines = ["BOP Part No. :\n"]
    for bop in bop_list: op10_lines.append(f"{bop['part_no']}-{bop['qty']}Nos")
    op10_text = "\n".join(op10_lines)
    
    op20_lines = ["BOP Part No. :\n"]
    for bop in bop_list: op20_lines.append(f"{bop['part_no']}-{bop['qty']}Nos")
    op20_text = "\n".join(op20_lines)

    wb = openpyxl.load_workbook(pfd_template)
    if BOP_SHEET_NAME not in wb.sheetnames:
        return {"status": "error", "message": "BOP sheet not found"}

    sheet = wb[BOP_SHEET_NAME]

    # Inject headers
    inject_sheet_headers(sheet, part_num, part_desc, rev)

    sym_cols = _get_process_symbol_columns(sheet)

    for row_num in range(8, 100): 
        val_a = sheet[f'A{row_num}'].value
        if val_a:
            cell_text = str(val_a).strip().upper()
            if "CFT TEAM MEMBERS" in cell_text or "REVISION HISTORY" in cell_text or "PREPARED BY" in cell_text:
                break
        
        # Check column C too
        val_c = str(sheet[f'C{row_num}'].value).strip().upper() if sheet[f'C{row_num}'].value else ""
        if "CFT TEAM MEMBERS" in val_c or "REVISION HISTORY" in val_c or "PREPARED BY" in val_c:
            break
        
        op_cell = str(val_a).strip() if val_a else ""
        
        if op_cell == "10.0" or op_cell == "10":
            safe_write_data(sheet, f'{BOP_TARGET_COL}{row_num}', op10_text)
        elif op_cell == "20.0" or op_cell == "20":
            safe_write_data(sheet, f'{BOP_TARGET_COL}{row_num}', op20_text)

        # Inject process symbol
        val_f = sheet[f'F{row_num}'].value
        desc_text = str(val_f) if val_f else ""
        if desc_text.strip():
            proc_type = _classify_row_process(desc_text)
            sym_col = sym_cols.get(proc_type)
            if sym_col:
                col_letter = openpyxl.utils.get_column_letter(sym_col)
                cell_ref = f"{col_letter}{row_num}"
                safe_write_data(sheet, cell_ref, SYMBOLS[proc_type], font=GREEN_FONT)

    wb.save(pfd_output)
    return {"status": "success", "bop_path": pfd_output}
