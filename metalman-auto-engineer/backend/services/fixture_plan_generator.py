"""
Fixture PM Master Plan Generator

Uses the "Fresh Workbook Transfuser" approach:
  1. Extract fixture records from the feasibility Excel
  2. Clone the template header (rows 1-8) exactly -- column widths, styles, merges
  3. Build data rows using style baseline from template row 9
  4. Transfuse the footer with precise row-offset shifting
"""

import os
import math
import openpyxl
from copy import copy
from openpyxl.styles import Alignment, Font, Border, Side


def generate_fixture_plan(
    feasibility_path,
    template_path,
    output_path,
    month_label=None,   # kept for API compatibility (unused in this approach)
):
    """
    Public API -- called by routes.py.

    Args:
        feasibility_path: Path to the feasibility xlsx.
        template_path:    Path to 'fixture_plan_template.xlsx' in assets.
        output_path:      Destination path for the generated xlsx.
        month_label:      Unused (template already contains the month cell).

    Returns:
        True on success, False on failure.
    """
    try:
        _generate_transfused_fixture_pm_plan(
            feasibility_xlsx=feasibility_path,
            template_xlsx=template_path,
            output_xlsx=output_path,
        )
        return True
    except Exception as e:
        print("[FP] ERROR: " + str(e), flush=True)
        import traceback
        traceback.print_exc()
        return False


# -----------------------------------------------------------------------------
# CORE ENGINE  (original script logic -- no logic changes, emojis removed)
# -----------------------------------------------------------------------------

def _generate_transfused_fixture_pm_plan(feasibility_xlsx, template_xlsx, output_xlsx):
    print("[FP] Initializing Fresh Workbook Transfuser Engine...", flush=True)

    # ==========================================
    # 1. EXTRACT DATA FROM FEASIBILITY
    # ==========================================
    src_wb = openpyxl.load_workbook(feasibility_xlsx, data_only=True)
    src_ws = src_wb.active

    extracted_data = []
    src_headers    = []
    header_row_idx = None

    # Auto-detect true data headers from feasibility sheet
    for r_idx in [3, 4, 5]:
        row_vals = [str(cell.value or "").strip().lower() for cell in src_ws[r_idx]]
        if any("fixture no" in v for v in row_vals):
            header_row_idx = r_idx
            src_headers    = row_vals
            break

    if not header_row_idx:
        raise ValueError("[FP] Could not locate structural header matrix in feasibility sheet rows 3-5.")

    # Column maps with automated defaults
    idx_part     = next((i for i, h in enumerate(src_headers) if h == "part name"),                16)
    idx_fix_no   = next((i for i, h in enumerate(src_headers) if "fixture no"   in h),            11)
    idx_op       = next((i for i, h in enumerate(src_headers) if "operation no" in h or h == "op no"), 1)
    idx_fix_name = next((i for i, h in enumerate(src_headers) if "fixture name" in h),            12)

    for r in range(header_row_idx + 1, src_ws.max_row + 1):
        part_name    = src_ws.cell(row=r, column=idx_part     + 1).value
        fixture_no   = src_ws.cell(row=r, column=idx_fix_no   + 1).value
        op_no        = src_ws.cell(row=r, column=idx_op       + 1).value
        fixture_name = src_ws.cell(row=r, column=idx_fix_name + 1).value

        f_no_str = str(fixture_no or "").strip().upper()
        if fixture_no and f_no_str not in ["NA", "N/A", "", "NONE"]:
            extracted_data.append({
                "part_name":    part_name    if part_name    else "DOOR, ASSY, LH",
                "fixture_no":   fixture_no,
                "op_no":        op_no,
                "fixture_name": fixture_name if fixture_name else "Welding Fixture",
            })

    print("[FP] Successfully extracted " + str(len(extracted_data)) + " operational machine data blocks.", flush=True)

    # ==========================================
    # 2. INITIALIZE CLEAN SEPARATE WORKBOOK
    # ==========================================
    template_wb = openpyxl.load_workbook(template_xlsx)
    template_ws = template_wb.active

    output_wb       = openpyxl.Workbook()
    output_ws       = output_wb.active
    output_ws.title = template_ws.title

    # Helper: replicate cell profiles safely
    def clone_cell_profile(source_cell, target_cell):
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font          = copy(source_cell.font)
            target_cell.fill          = copy(source_cell.fill)
            target_cell.border        = copy(source_cell.border)
            target_cell.alignment     = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format

    # TRANSFUSION STEP 1: Copy exact column configurations / widths
    for col_idx in range(1, 40): # Scan broad range of columns
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        width = template_ws.column_dimensions[col_letter].width
        if width:
            output_ws.column_dimensions[col_letter].width = width
        else:
            output_ws.column_dimensions[col_letter].width = 12 # Default visibility

    # ==========================================
    # 3. TRANSFUSE HEADER MATRIX (ROWS 1 TO 8)
    # ==========================================
    print("[FP] Header Block Transfusion active...", flush=True)
    for r in range(1, 9):
        # FIX: Significantly increase header row heights (especially for logos and title)
        h = template_ws.row_dimensions[r].height
        if r == 1:
            output_ws.row_dimensions[r].height = 100 # Maximum space for Logo
        elif r <= 5:
            output_ws.row_dimensions[r].height = 45 # Space for Model/Line info
        else:
            output_ws.row_dimensions[r].height = max(30, h if h else 30)
            
        for c in range(1, 37):
            clone_cell_profile(template_ws.cell(row=r, column=c), output_ws.cell(row=r, column=c))

    # Transfuse native header merge spaces
    for merge_range in template_ws.merged_cells.ranges:
        if merge_range.max_row <= 8:
            output_ws.merge_cells(str(merge_range))

    # Transfuse images (such as the Metalman logo) from the template
    if hasattr(template_ws, "_images"):
        for img in template_ws._images:
            new_img = copy(img)
            new_img.anchor = img.anchor
            output_ws.add_image(new_img)

    # ==========================================
    # 4. CAPTURE DATA ZONE STYLE BASELINES
    # ==========================================
    style_baseline = {}
    for c in range(1, 37):
        base_cell = template_ws.cell(row=9, column=c)
        style_baseline[c] = {
            'font':          copy(base_cell.font),
            'fill':          copy(base_cell.fill),
            'border':        copy(base_cell.border),
            'number_format': base_cell.number_format,
        }

    # ==========================================
    # 5. DYNAMICALLY BUILD DATA REGISTRY ZONE
    # ==========================================
    print("[FP] Injecting and formatting production registry rows...", flush=True)
    current_row = 9

    for idx, item in enumerate(extracted_data, start=1):
        str_part     = str(item['part_name']    or "")
        str_fix_name = str(item['fixture_name'] or "")
        str_fix_no   = str(item['fixture_no']   or "")

        # Calculate text breaks based on column layouts
        lines_part      = math.ceil(len(str_part)     / 24)
        lines_fix_name  = math.ceil(len(str_fix_name) / 28)
        lines_fix_no    = math.ceil(len(str_fix_no)   / 22)
        explicit_breaks = max(
            str_part.count('\n'),
            str_fix_name.count('\n'),
            str_fix_no.count('\n'),
        ) + 1

        max_lines = max(lines_part, lines_fix_name, lines_fix_no, explicit_breaks)
        output_ws.row_dimensions[current_row].height = max(26, (max_lines * 14) + 12)

        # Populate raw values
        output_ws.cell(row=current_row, column=1, value=idx)
        output_ws.cell(row=current_row, column=2, value=item['part_name'])
        output_ws.cell(row=current_row, column=3, value=item['fixture_no'])
        output_ws.cell(row=current_row, column=4, value=item['op_no'])
        output_ws.cell(row=current_row, column=5, value=item['fixture_name'])

        # Apply style baseline + alignment to every data column
        for c in range(1, 37):
            cell = output_ws.cell(row=current_row, column=c)
            base = style_baseline[c]
            if base['font']:   cell.font   = base['font']
            if base['fill']:   cell.fill   = base['fill']
            if base['border']: cell.border = base['border']
            cell.number_format = base['number_format']

            # Unified alignment with explicit auto-wrapping
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        current_row += 1

    # ==========================================
    # 6. IDENTIFY & TRANSFUSE FOOTER MATRIX
    # ==========================================
    print("[FP] Appending and locking shifting footers...", flush=True)
    template_footer_start = None
    for r in range(9, template_ws.max_row + 1):
        val    = str(template_ws.cell(row=r, column=1).value or "").strip().upper()
        val_c5 = str(template_ws.cell(row=r, column=5).value or "").strip().upper()
        if "PLAN-" in val or "CHECKED BY" in val_c5 or "PLAN ACHIEVED-" in val:
            template_footer_start = r
            break

    if not template_footer_start:
        template_footer_start = template_ws.max_row - 5

    # Identify TRUE template data end (to avoid ghost rows in footer)
    template_footer_end = template_ws.max_row
    try:
        # Find the last row in the footer block that has any value
        res = template_ws.cell(row=template_ws.max_row, column=1).row
        for r in range(template_ws.max_row, template_footer_start, -1):
            if any(template_ws.cell(r, c).value for c in range(1, 10)): # Check first few columns
                template_footer_end = r
                break
    except: pass

    row_offset = current_row - template_footer_start

    for r in range(template_footer_start, template_footer_end + 1):
        target_r = r + row_offset
        output_ws.row_dimensions[target_r].height = template_ws.row_dimensions[r].height
        for c in range(1, 37):
            clone_cell_profile(template_ws.cell(row=r, column=c), output_ws.cell(row=target_r, column=c))

    # Transfuse & shift template footer cell merges precisely
    for merge_range in template_ws.merged_cells.ranges:
        if merge_range.min_row >= template_footer_start:
            shifted_range = openpyxl.worksheet.cell_range.CellRange(
                min_col=merge_range.min_col,
                max_col=merge_range.max_col,
                min_row=merge_range.min_row + row_offset,
                max_row=merge_range.max_row + row_offset,
            )
            output_ws.merge_cells(str(shifted_range))

    # Enforce underlying gridline visibility
    output_ws.views.sheetView[0].showGridLines = True

    output_wb.save(output_xlsx)
    print("[FP] MASTER PLAN TRANSFUSION COMPLETE: " + output_xlsx, flush=True)
