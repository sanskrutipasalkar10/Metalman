import json
import re
import copy
import os
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG & CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
OP_FILL   = PatternFill("solid", fgColor="92D050")
OP_FONT   = Font(bold=True, size=28)
OP_ALIGN  = Alignment(horizontal="center", vertical="center", indent=1)

NO_FONT      = Font(bold=True, size=24)
NO_ALIGN     = Alignment(horizontal="center", vertical="center")
FUNC_FONT    = Font(size=16)
FUNC_ALIGN   = Alignment(wrap_text=True, vertical="center")
NORMAL_FONT  = Font(size=16)
NORMAL_ALIGN = Alignment(wrap_text=True, vertical="center")
MOVE_LABEL_FONT = Font(bold=False, size=24)

THIN  = Side(border_style="thin")
MED   = Side(border_style="medium")
NONE_ = Side(border_style=None)

MAX_COL = 19   # columns A-S

# BOP/SM Specific Constants
COL_WIDTHS = {
    7:  57,   # G – Cause
    9:  69,   # I – Prevention
    10: 61,   # J – Detection
    4:  35,   # D – Effect
    3:  29,   # C – Failure mode
    2:  28,   # B – Process fn
}
LINE_HEIGHT_PT = 14
MIN_ROW_PT_BOP = 48
MIN_ROW_PT_SM  = 18

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def clone_font(cell):
    f = cell.font
    return Font(name=f.name, bold=f.bold, italic=f.italic, size=f.size,
                color=copy.copy(f.color))

def clone_fill(cell):
    fl = cell.fill
    if fl.fill_type in (None, "none"):
        return PatternFill()
    return PatternFill(fill_type=fl.fill_type,
                       fgColor=copy.copy(fl.fgColor),
                       bgColor=copy.copy(fl.bgColor))

def clone_border(cell):
    b = cell.border
    def side(s): return Side(border_style=s.border_style, color=copy.copy(s.color))
    return Border(left=side(b.left), right=side(b.right),
                  top=side(b.top), bottom=side(b.bottom))

def clone_alignment(cell):
    a = cell.alignment
    return Alignment(horizontal=a.horizontal, vertical=a.vertical,
                     wrap_text=a.wrap_text, shrink_to_fit=a.shrink_to_fit,
                     indent=a.indent)

def copy_cell_style(src_cell, dst_cell):
    dst_cell.font = clone_font(src_cell)
    dst_cell.fill = clone_fill(src_cell)
    dst_cell.border = clone_border(src_cell)
    dst_cell.alignment = clone_alignment(src_cell)
    dst_cell.number_format = src_cell.number_format

def data_border(col_letter):
    left = MED if col_letter == "A" else THIN
    return Border(left=left, right=THIN, top=THIN, bottom=THIN)

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def style_cell_bop(cell, value=None, bold=False, size=11,
               horizontal="left", vertical="center", wrap_text=True, border=True):
    if value is not None:
        cell.value = value
    cell.font      = Font(name="Arial", bold=bold, size=size)
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
    if border:
        cell.border = _thin_border()

def style_empty_bop(cell):
    cell.value     = None
    cell.font      = Font(name="Arial", size=11)
    cell.alignment = Alignment(horizontal="center", wrap_text=False)
    cell.border    = _thin_border()

# ─────────────────────────────────────────────────────────────────────────────
# DATA HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def normalise(text: str) -> str:
    if not isinstance(text, str): return ""
    return re.sub(r"[^A-Z0-9]+", "_", text.upper()).strip("_")

def index_pfmea_data(data):
    indexed = {}
    if isinstance(data, dict):
        for k, v in data.items():
            indexed[normalise(k)] = v
    elif isinstance(data, list):
        for item in data:
            op_key = item.get("Operation") or item.get("id") or item.get("_operation_title")
            if op_key:
                indexed[normalise(op_key)] = item
    
    if "FINISHING_DEBURRING" in indexed:
        indexed["FINSHING_DEBURRING"] = indexed["FINISHING_DEBURRING"]
    return indexed

def lookup_function(func_name: str, pfmea_indexed: dict):
    n = normalise(func_name)
    if n in pfmea_indexed:
        return pfmea_indexed[n]
    
    matches = [(k, v) for k, v in pfmea_indexed.items() if k in n or n in k]
    if matches:
        best = max(matches, key=lambda x: len(x[0]))
        return best[1]
        
    for k, v in pfmea_indexed.items():
        tokens = [t for t in k.split("_") if len(t) > 3]
        if tokens and all(t in n for t in tokens):
            return v
    return None

def format_effects(effects_val):
    if isinstance(effects_val, dict):
        parts = [f"{k}{v}" for k, v in effects_val.items()]
        return "\n".join(parts)
    return str(effects_val) if effects_val else ""

def text_height(text, col_width, min_h=18):
    if not text:
        return min_h
    effective = col_width * 0.85
    lines = str(text).split("\n")
    total_lines = 0
    for line in lines:
        if not line.strip():
            total_lines += 1
        else:
            total_lines += max(1, math.ceil(len(line) / effective))
    return max(min_h, total_lines * LINE_HEIGHT_PT + 10)

def merged_block_height_sm(rows_data, effect_str, fm_str):
    base_heights = []
    for r in rows_data:
        h = max(
            text_height(r.get("Potential Cause(s) of Failure", ""), COL_WIDTHS[7], MIN_ROW_PT_SM),
            text_height(r.get("Current Controls Prevention", ""),    COL_WIDTHS[9], MIN_ROW_PT_SM),
            text_height(r.get("Current Controls Detection", ""),     COL_WIDTHS[10], MIN_ROW_PT_SM)
        )
        base_heights.append(h)
 
    merged_h = max(
        text_height(fm_str, COL_WIDTHS[3], MIN_ROW_PT_SM),
        text_height(effect_str, COL_WIDTHS[4], MIN_ROW_PT_SM)
    )
 
    current_total = sum(base_heights)
    if current_total < merged_h:
        deficit = merged_h - current_total
        extra = deficit / len(base_heights)
        base_heights = [h + extra for h in base_heights]
 
    return base_heights

# ─────────────────────────────────────────────────────────────────────────────
# WRITING HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def write_op_header(ws, row_num, op_label, fixture_name):
    ws.row_dimensions[row_num].height = 45.75
    for c in range(1, MAX_COL + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill  = PatternFill("solid", fgColor="92D050")
        cell.font  = OP_FONT
        cell.alignment = OP_ALIGN
        cell.border = Border(left=MED, right=MED, top=MED, bottom=MED)
    
    ws.cell(row=row_num, column=1).value = op_label
    ws.cell(row=row_num, column=1).alignment = OP_ALIGN
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    
    ws.cell(row=row_num, column=7).value = fixture_name
    ws.cell(row=row_num, column=7).alignment = OP_ALIGN
    ws.merge_cells(start_row=row_num, start_column=7, end_row=row_num, end_column=MAX_COL)

def write_data_row(ws, row_num, seq_no, func_name, failure_mode,
                   effects_text, severity, cls, causes,
                   occurrence, prevention, detection, det_val):
    ws.row_dimensions[row_num].height = None
    values = {
        1: seq_no, 2: func_name, 3: failure_mode, 4: effects_text,
        5: severity, 6: cls, 7: causes, 8: occurrence,
        9: prevention, 10: detection, 11: det_val,
        12: f"=K{row_num}*H{row_num}*E{row_num}",
    }
    for c in range(1, MAX_COL + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.value = values.get(c, None)
        cell.border = data_border(cell.column_letter)
        if c == 1:
            cell.font = NO_FONT; cell.alignment = NO_ALIGN
        elif c == 2:
            cell.font = FUNC_FONT; cell.alignment = FUNC_ALIGN
        else:
            cell.font = NORMAL_FONT; cell.alignment = NORMAL_ALIGN

def write_move_row(ws, row_num, seq_no, label):
    for c in range(1, MAX_COL + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.border = data_border(cell.column_letter)
        cell.font = NO_FONT if c == 1 else MOVE_LABEL_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_num, column=1).value = seq_no
    ws.cell(row=row_num, column=2).value = label
    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=MAX_COL)

def safe_merge(ws, r1, c1, r2, c2):
    if r2 == r1 and c2 == c1: return
    try: ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    except: pass

# ─────────────────────────────────────────────────────────────────────────────
# SHEET PROCESSORS
# ─────────────────────────────────────────────────────────────────────────────
def parse_feasibility_assembly(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    op_col = func_col = fixture_col = header_row = None
    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value) if cell.value else ""
            if "Operation No" in v:
                header_row = cell.row
                op_col = cell.column
            if "Process Function" in v and header_row == cell.row:
                func_col = cell.column
            if "Fixture Name" in v and header_row == cell.row:
                fixture_col = cell.column
        if op_col and func_col and fixture_col: break
    if not (op_col and func_col and fixture_col): return []
    operations = []
    current_op = current_fixture = None
    for row in ws.iter_rows(min_row=header_row + 2):
        rn = row[0].row
        op_val  = ws.cell(row=rn, column=op_col).value
        func_val = ws.cell(row=rn, column=func_col).value
        fix_val  = ws.cell(row=rn, column=fixture_col).value
        op_str = str(op_val).strip() if op_val else ""
        if op_str.upper().startswith("OP-"):
            current_op = op_str.replace(" ", "")
            current_fixture = str(fix_val).strip() if fix_val else ""
        if func_val and current_op:
            raw_functions = str(func_val).replace("\n", " ")
            func_list = [f.strip() for f in raw_functions.split(",") if f.strip()]
            operations.append({"op": current_op, "fixture": current_fixture, "functions": func_list})
            current_op = None
    return operations

def process_generic_sheet(ws_out, ws_src, pfmea_dict, min_row_pt):
    # 1. Header
    HEADER_END = 10
    for col, dim in ws_src.column_dimensions.items():
        ws_out.column_dimensions[col].width = dim.width
    for r in range(1, HEADER_END + 1):
        if r in ws_src.row_dimensions:
            ws_out.row_dimensions[r].height = ws_src.row_dimensions[r].height
        for c in range(1, MAX_COL + 1):
            src = ws_src.cell(row=r, column=c)
            dst = ws_out.cell(row=r, column=c, value=src.value)
            copy_cell_style(src, dst)
    for m in ws_src.merged_cells.ranges:
        if m.max_row <= HEADER_END:
            ws_out.merge_cells(str(m))
    
    if hasattr(ws_src, "_images"):
        for img in ws_src._images:
            new_img = copy.copy(img)
            new_img.anchor = img.anchor
            ws_out.add_image(new_img)
    
    # 2. Blocks
    blocks = []
    for i, (proc_key, failure_modes) in enumerate(pfmea_dict.items()):
        step = (i + 1) * 10
        if not failure_modes:
            blocks.append({"step": step, "process_fn": proc_key.replace("_"," ").title(), "fm": None, "effect_str": None, "rows": []})
            continue
        for fm_key, cause_rows in failure_modes.items():
            if not cause_rows: continue
            eff = cause_rows[0].get("Potential Effect of failure", {})
            if isinstance(eff, dict):
                effect_str = f"\nM/C:-:{eff.get('M/C:-:','')}\nSO -: {eff.get('SO -:','')}\nCU:- {eff.get('CU:-','')}\nEU- {eff.get('EU-','')}\n "
            else:
                effect_str = str(eff)
            blocks.append({"step": step, "process_fn": cause_rows[0].get("Process function", proc_key), "fm": fm_key, "effect_str": effect_str, "rows": cause_rows})
    
    DATA_START = HEADER_END + 1
    cur = DATA_START
    step_spans = {}

    for blk in blocks:
        step = blk["step"]; proc_fn = blk["process_fn"]; fm = blk["fm"]; eff_str = blk["effect_str"]; rows = blk["rows"]
        if fm is None:
            ws_out.row_dimensions[cur].height = 48
            style_cell_bop(ws_out.cell(cur,1), step, bold=True, size=20, horizontal="center")
            style_cell_bop(ws_out.cell(cur,2), "Move for Next Process", size=14, horizontal="center", vertical="center", wrap_text=True)
            safe_merge(ws_out, cur, 2, cur, 19)
            for c in range(3, 20): ws_out.cell(cur,c).border = _thin_border()
            step_spans.setdefault(step, [cur, cur])[1] = cur
            cur += 1; continue

        n = len(rows); blk_start = cur; blk_end = cur + n - 1
        step_spans.setdefault(step, [blk_start, blk_end])[1] = blk_end
        style_cell_bop(ws_out.cell(blk_start,1), step, bold=True, size=20, horizontal="center")
        for r_fn in range(blk_start, blk_end + 1):
            style_cell_bop(ws_out.cell(r_fn, 2), proc_fn, size=14, horizontal="center", vertical="center", wrap_text=True)
        style_cell_bop(ws_out.cell(blk_start,3), fm, size=11, horizontal="left", vertical="center", wrap_text=True)
        if n > 1: safe_merge(ws_out, blk_start, 3, blk_end, 3)
        style_cell_bop(ws_out.cell(blk_start,4), eff_str, size=11, horizontal="left", vertical="center", wrap_text=True)
        if n > 1: safe_merge(ws_out, blk_start, 4, blk_end, 4)

        row_heights = merged_block_height_sm(rows, eff_str, fm) if min_row_pt == MIN_ROW_PT_SM else [text_height(r.get("Potential Cause(s) of Failure", ""), COL_WIDTHS[7], min_row_pt) for r in rows]

        for i, cr in enumerate(rows):
            r = blk_start + i
            ws_out.row_dimensions[r].height = row_heights[i]
            style_empty_bop(ws_out.cell(r, 5)); style_empty_bop(ws_out.cell(r, 6))
            style_cell_bop(ws_out.cell(r,7), cr.get("Potential Cause(s) of Failure", ""), size=11, horizontal="left", wrap_text=True)
            style_empty_bop(ws_out.cell(r, 8))
            style_cell_bop(ws_out.cell(r,9), cr.get("Current Controls Prevention", ""), size=11, horizontal="left", wrap_text=True)
            style_cell_bop(ws_out.cell(r,10), cr.get("Current Controls Detection", ""), size=11, horizontal="left", wrap_text=True)
            style_empty_bop(ws_out.cell(r, 11))
            ws_out.cell(r,12).value = f"=IF(OR(E{r}=\"\",H{r}=\"\",K{r}=\"\"),\"\",E{r}*H{r}*K{r})"
            ws_out.cell(r,12).font = Font(name="Arial", size=11); ws_out.cell(r,12).alignment = Alignment(horizontal="center", wrap_text=False); ws_out.cell(r,12).border = _thin_border()
            for c in range(13, 20): style_empty_bop(ws_out.cell(r, c))
            if i > 0:
                for c in (1, 2, 3, 4): ws_out.cell(r, c).border = _thin_border()
        cur = blk_end + 1

    for step, (rs, re) in step_spans.items():
        if re > rs: safe_merge(ws_out, rs, 1, re, 1)

    # Merge col B per same-process-fn
    grp_start = DATA_START; prev_fn = ws_out.cell(DATA_START, 2).value
    for r in range(DATA_START + 1, cur):
        fn = ws_out.cell(r, 2).value
        if fn != prev_fn:
            if (r - 1) > grp_start: safe_merge(ws_out, grp_start, 2, r - 1, 2)
            grp_start = r; prev_fn = fn
    if (cur - 1) > grp_start: safe_merge(ws_out, grp_start, 2, cur - 1, 2)

    # Footer
    abbrev_row = None
    for row in ws_src.iter_rows():
        for cell in row:
            if cell.value and "ABBREVIATION" in str(cell.value).upper(): abbrev_row = cell.row; break
        if abbrev_row: break
    if abbrev_row:
        offset = cur - abbrev_row
        for orig_r in range(abbrev_row, ws_src.max_row + 1):
            tgt_r = orig_r + offset
            ws_out.row_dimensions[tgt_r].height = ws_src.row_dimensions[orig_r].height or 30
            for orig_cell in ws_src[orig_r]:
                tgt_cell = ws_out.cell(tgt_r, orig_cell.column)
                tgt_cell.value = orig_cell.value
                copy_cell_style(orig_cell, tgt_cell)
        for m in ws_src.merged_cells.ranges:
            if m.min_row >= abbrev_row:
                safe_merge(ws_out, m.min_row + offset, m.min_col, m.max_row + offset, m.max_col)

def process_assembly_sheet(ws_out, ws_src, feasibility_path, dict_path):
    if not os.path.exists(dict_path): return
    with open(dict_path, encoding="utf-8") as fh: pfmea_raw = json.load(fh)
    pfmea_indexed = index_pfmea_data(pfmea_raw)
    ops_data = parse_feasibility_assembly(feasibility_path)
    
    # Header
    DATA_START_ROW = None
    for row in ws_src.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip().upper().startswith("OP-"): DATA_START_ROW = cell.row; break
        if DATA_START_ROW: break
    HEADER_ROWS = (DATA_START_ROW - 1) if DATA_START_ROW else 9
    for col, dim in ws_src.column_dimensions.items(): ws_out.column_dimensions[col].width = dim.width
    for r in range(1, HEADER_ROWS + 1):
        if r in ws_src.row_dimensions: ws_out.row_dimensions[r].height = ws_src.row_dimensions[r].height
        for c in range(1, MAX_COL + 1):
            src_cell = ws_src.cell(row=r, column=c); dst = ws_out.cell(row=r, column=c, value=src_cell.value)
            copy_cell_style(src_cell, dst)
            val_str = str(dst.value).strip() if dst.value else ""
            if r == 1: dst.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif c >= 15 or val_str.endswith(":-") or val_str.endswith(":"): dst.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            elif (2 <= r <= 5) and (3 <= c <= 14): dst.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            elif r >= 6: dst.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else: dst.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    for mr in ws_src.merged_cells.ranges:
        if mr.min_row <= HEADER_ROWS: ws_out.merge_cells(str(mr))
    if hasattr(ws_src, "_images"):
        for img in ws_src._images:
            new_img = copy.copy(img); new_img.anchor = img.anchor; ws_out.add_image(new_img)

    # Operations
    current_row = HEADER_ROWS + 1
    for op_info in ops_data:
        op_label = op_info["op"]; fixture = op_info["fixture"]; functions = op_info["functions"]
        write_op_header(ws_out, current_row, op_label, fixture); current_row += 1
        step_no = 10
        for func_name in functions:
            entries = lookup_function(func_name, pfmea_indexed)
            if entries is None: continue
            func_start_row = current_row
            for f_idx, (failure_mode, rows) in enumerate(entries.items()):
                fm_start_row = current_row
                for r_idx, row_data in enumerate(rows):
                    causes = row_data.get("Potential Cause(s) of Failure", ""); prevention = row_data.get("Current Controls Prevention", ""); detection = row_data.get("Current Controls Detection", "")
                    effects = row_data.get("Potential Effect of failure", ""); cls = row_data.get("Class", ""); effects_str = format_effects(effects)
                    write_data_row(ws_out, current_row, seq_no=step_no if (f_idx == 0 and r_idx == 0) else None, func_name=func_name if (f_idx == 0 and r_idx == 0) else None,
                                   failure_mode=failure_mode if (r_idx == 0) else None, effects_text=effects_str if (r_idx == 0) else None, severity="", cls=cls, causes=causes,
                                   occurrence="", prevention=prevention, detection=detection, det_val="")
                    current_row += 1
                if current_row - 1 > fm_start_row:
                    for c in (3, 4, 5, 6): ws_out.merge_cells(start_row=fm_start_row, start_column=c, end_row=current_row-1, end_column=c)
            if current_row - 1 > func_start_row:
                for c in (1, 2): ws_out.merge_cells(start_row=func_start_row, start_column=c, end_row=current_row-1, end_column=c)
            step_no += 10
        last_norm = normalise(functions[-1]) if functions else ""
        if "MOVE_" not in last_norm and "DISPATCH" not in last_norm:
            footer_label = "Move to Dispatch" if op_info == ops_data[-1] else "Move for Next Process"
            write_move_row(ws_out, current_row, step_no, footer_label); current_row += 1

    # Footer
    abbrev_row = None
    for r in range(10, ws_src.max_row + 1):
        val = ws_src.cell(row=r, column=1).value
        if val and "ABBREVIATION" in str(val).upper(): abbrev_row = r; break
    if abbrev_row:
        for r in range(abbrev_row, ws_src.max_row + 1):
            for c in range(1, MAX_COL + 1):
                src_cell = ws_src.cell(row=r, column=c); dst_cell = ws_out.cell(row=current_row, column=c, value=src_cell.value); copy_cell_style(src_cell, dst_cell)
            for merge in ws_src.merged_cells.ranges:
                if merge.min_row == r: ws_out.merge_cells(start_row=current_row, start_column=merge.min_col, end_row=current_row + (merge.max_row - merge.min_row), end_column=merge.max_col)
            current_row += 1

def generate_pfmea_excel(feasibility_path, template_path, output_path, dict_dir):
    wb_tmpl = openpyxl.load_workbook(template_path)
    wb_out = openpyxl.Workbook()
    default_sheet = wb_out.active; wb_out.remove(default_sheet)

    # 1. Assembly
    if "Assembly PFMEA" in wb_tmpl.sheetnames:
        ws_asm_out = wb_out.create_sheet("Assembly PFMEA"); ws_asm_src = wb_tmpl["Assembly PFMEA"]
        asm_dict_path = os.path.join(dict_dir, "assembly", "pfmea_assembly_dict_op2.json")
        process_assembly_sheet(ws_asm_out, ws_asm_src, feasibility_path, asm_dict_path)
    
    # 2. Sheetmetal
    if "Sheet Metal Parts PFMEA" in wb_tmpl.sheetnames:
        ws_sm_out = wb_out.create_sheet("Sheet Metal Parts PFMEA"); ws_sm_src = wb_tmpl["Sheet Metal Parts PFMEA"]
        sm_dict_path = os.path.join(dict_dir, "sheetmetal", "pfmea_sheet_metal_part_dict.json")
        if os.path.exists(sm_dict_path):
            with open(sm_dict_path, encoding="utf-8") as f: sm_dict = json.load(f)
            process_generic_sheet(ws_sm_out, ws_sm_src, sm_dict, MIN_ROW_PT_SM)
    
    # 3. BOP
    if "BOP Parts PFMEA" in wb_tmpl.sheetnames:
        ws_bop_out = wb_out.create_sheet("BOP Parts PFMEA"); ws_bop_src = wb_tmpl["BOP Parts PFMEA"]
        bop_dict_path = os.path.join(dict_dir, "bop", "pfmea_bop_dict.json")
        if os.path.exists(bop_dict_path):
            with open(bop_dict_path, encoding="utf-8") as f: bop_dict = json.load(f)
            process_generic_sheet(ws_bop_out, ws_bop_src, bop_dict, MIN_ROW_PT_BOP)

    wb_out.save(output_path); return True
