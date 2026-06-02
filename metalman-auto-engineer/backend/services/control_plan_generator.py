"""
Control Plan Document Generator
Handles all 4 sheets:
  1. Sub Assmbly Index     – reuses feasibility data (same as PFD)
  2. ASSY_SUB_ASSY          – placeholder (user will provide code)
  3. Q-Sheetmetal Parts     – dynamic: feasibility + PFD output + dict
  4. T - BOP & Hardwares    – static dict injection
"""
import os, json, copy, warnings, re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

# ─── SHEET NAMES ──────────────────────────────────────────────────────────────
SUB_ASSY_SHEET  = "Sub Assmbly Index"
ASSY_ASSY_SHEET = "ASSY_SUB_ASSY"
Q_SM_SHEET      = "Q-Sheetmetal Parts"
BOP_SHEET       = "T - BOP & Hardwares"

# ─── COLOURS ──────────────────────────────────────────────────────────────────
GREEN_PART = "73FDC5"
GREEN_ASSY = "76FAB2"
GREEN_PAINT = "83F37D"

# ─── STYLE HELPERS ────────────────────────────────────────────────────────────
def _s(style="thin", color="000000"): return Side(border_style=style, color=color)
def _border(left="thin", right="thin"):
    l_style = None if left == "none" else (left if left else "thin")
    r_style = None if right == "none" else (right if right else "thin")
    l = Side(border_style=l_style, color="000000") if l_style else Side(border_style=None)
    r = Side(border_style=r_style, color="000000") if r_style else Side(border_style=None)
    return Border(left=l, right=r, top=_s(), bottom=_s())
def _std_border():
    s = _s()
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="center", v="center", wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _font(bold=False, size=11, name="Arial"): return Font(name=name, bold=bold, size=size)
def _fill(rgb): return PatternFill("solid", fgColor=rgb)

def sc(cell, value=None, bold=False, size=11, rgb=None, h="center", v="center", wrap=True, left="thin", right="thin"):
    if value is not None:
        try:
            cell.value = value
        except AttributeError:
            pass
    cell.font      = _font(bold, size)
    cell.alignment = _align(h, v, wrap)
    cell.border    = _border(left, right)
    if rgb:
        cell.fill = _fill(rgb)

def _copy_cell(src, dst):
    try:
        dst.value = src.value
    except AttributeError:
        pass
    if src.has_style:
        dst.font         = copy.copy(src.font)
        dst.fill         = copy.copy(src.fill)
        dst.alignment    = copy.copy(src.alignment)
        dst.border       = copy.copy(src.border)
        dst.number_format = src.number_format

def _safe_merge(ws, r1, c1, r2, c2):
    if r1 == r2 and c1 == c2:
        return
    try:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    except Exception:
        pass

def _clear_below_header(ws, header_end_row):
    # Unmerge all cells below header_end_row
    merged_ranges = list(ws.merged_cells.ranges)
    for r in merged_ranges:
        if r.min_row > header_end_row:
            try:
                ws.unmerge_cells(str(r))
            except Exception:
                pass
    # Delete all rows below header_end_row
    if ws.max_row > header_end_row:
        ws.delete_rows(header_end_row + 1, ws.max_row - header_end_row)

# ─── SHEET 1: SUB ASSY INDEX ──────────────────────────────────────────────────
def process_sub_assy_index_sheet(ws, assy_data, img_dir=""):
    """
    Fill Sub Assmbly Index sheet from feasibility assy_data list.
    assy_data: list of dicts {index, operation, part_name, part_no, rev}
    """
    if not assy_data:
        return
    start_row = 3
    
    # Colors for alternating patches
    light_blue = PatternFill("solid", fgColor="DCE6F1")
    dark_blue  = PatternFill("solid", fgColor="B8CCE4")
    font_fixed = Font(name="Arial", size=11)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, item in enumerate(assy_data):
        r = start_row + idx
        fill = light_blue if idx % 2 == 0 else dark_blue
        
        # Apply formatting to columns 1-6
        for c in range(1, 7):
            cell = ws.cell(r, c)
            cell.fill = fill
            cell.font = font_fixed
            cell.alignment = align_center
            cell.border = _std_border()

        ws.cell(r, 1).value = idx + 1
        ws.cell(r, 2).value = item.get("operation", "")
        ws.cell(r, 3).value = item.get("part_name", "")
        ws.cell(r, 4).value = item.get("part_no", "")
        ws.cell(r, 5).value = item.get("rev", "")
        
        # Image
        if img_dir:
            op_safe = str(item.get("operation", "")).replace(" ", "_").replace("/", "_")
            for fname in [f"{op_safe}.png", f"{item.get('part_no','')}.png"]:
                img_path = os.path.join(img_dir, fname)
                if os.path.exists(img_path):
                    try:
                        img = XLImage(img_path)
                        ws.row_dimensions[r].height = 140
                        img.width, img.height = 180, 130
                        # Center image in Col F (Col 6)
                        ws.add_image(img, ws.cell(r, 6).coordinate)
                    except Exception:
                        pass
                    break
            else:
                ws.row_dimensions[r].height = 25
        else:
            ws.row_dimensions[r].height = 25

# ─── SHEET 2: ASSY_SUB_ASSY ───────────────────────────────────────────────────
def _normalize(s):
    s = str(s).lower().strip()
    s = re.sub(r'[\s\-_/\\]+', ' ', s)
    return s

def _match_json_key(op_desc, cp_dict):
    needle = _normalize(op_desc)
    for key in cp_dict:
        if _normalize(key) == needle:
            return cp_dict[key]
    for key in cp_dict:
        k = _normalize(key)
        if k in needle or needle in k:
            return cp_dict[key]
    needle_words = set(needle.split())
    best_score, best_val = 0, None
    for key in cp_dict:
        kw = set(_normalize(key).split())
        score = len(needle_words & kw)
        if score > best_score:
            best_score = score
            best_val = cp_dict[key]
    return best_val if best_score >= 2 else None

def _load_pfd_ops(pfd_path, sheet_name):
    groups = []
    current_group = None
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(pfd_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"[CP] WARNING: PFD sheet '{sheet_name}' not found")
        return groups
    ws = wb[sheet_name]
    for r in range(8, ws.max_row + 1):
        col_a = str(ws.cell(r, 1).value or "").strip()
        col_f = str(ws.cell(r, 6).value or "").strip()
        col_g = str(ws.cell(r, 7).value or "").strip()
        col_h = str(ws.cell(r, 8).value or "").strip()
        if col_a.upper().startswith("OP-"):
            current_group = {"op_label": col_a, "parts_list": col_h, "operations": []}
            groups.append(current_group)
            continue
        if col_a.isdigit() and current_group is not None:
            machine = col_g if col_g not in ("N/A", "nan", "None", "") else ""
            current_group["operations"].append({
                "op_no": col_a, "op_desc": col_f, "machine": machine, "parts": col_h
            })
    return groups

def process_assy_sub_assy_sheet(ws_out, tmpl_ws, pfd_path, cp_dict):
    HEADER_END_ROW = 18
    _clear_below_header(ws_out, HEADER_END_ROW)
    FOOTER_NOTE_TEXT = "Note :"
    
    FOOTER_START = None
    for r in range(50, tmpl_ws.max_row + 1):
        if FOOTER_NOTE_TEXT in str(tmpl_ws.cell(r, 1).value or ""):
            FOOTER_START = r
            break
    FOOTER_END = tmpl_ws.max_row
    if not FOOTER_START:
        FOOTER_START = tmpl_ws.max_row - 20
        
    for col_ltr, dim in tmpl_ws.column_dimensions.items():
        if dim.width:
            ws_out.column_dimensions[col_ltr].width = dim.width
            
    for row in tmpl_ws.iter_rows(min_row=1, max_row=HEADER_END_ROW):
        ws_out.row_dimensions[row[0].row].height = tmpl_ws.row_dimensions[row[0].row].height or 15
        for src in row:
            _copy_cell(src, ws_out.cell(src.row, src.column))
    for mr in tmpl_ws.merged_cells.ranges:
        if mr.max_row <= HEADER_END_ROW:
            try: ws_out.merge_cells(str(mr))
            except Exception: pass
            
    op_groups = _load_pfd_ops(pfd_path, "SUB_ASSY")
    cur = HEADER_END_ROW + 1
    
    for group in op_groups:
        op_label = group["op_label"]
        parts_txt = group["parts_list"]
        ops = group["operations"]
        op_color = GREEN_PAINT if "OP-30" in op_label else GREEN_ASSY
        
        ws_out.row_dimensions[cur].height = 22.0
        sc(ws_out.cell(cur, 1), op_label, bold=True, size=12, rgb=op_color, left="medium", right="medium")
        for c in range(2, 9):
            sc(ws_out.cell(cur, c), rgb=op_color, left="none", right="none")
        try: ws_out.cell(cur, 2).value = "Sub Assembly"
        except AttributeError: pass
        ws_out.cell(cur, 2).font = _font(bold=True, size=12)
        ws_out.cell(cur, 2).border = _border(left="medium")
        _safe_merge(ws_out, cur, 2, cur, 8)
        
        for c in range(9, 15):
            sc(ws_out.cell(cur, c), rgb=op_color, h="left", left="none", right="none")
        try: ws_out.cell(cur, 9).value = parts_txt
        except AttributeError: pass
        ws_out.cell(cur, 9).font = _font(bold=True, size=11)
        ws_out.cell(cur, 9).alignment = _align(h="left")
        ws_out.cell(cur, 9).border = _border(left="medium")
        ws_out.cell(cur, 14).border = _s("medium", "000000") # right border
        _safe_merge(ws_out, cur, 9, cur, 14)
        cur += 1
        
        for op_idx, op in enumerate(ops):
            op_no = op["op_no"]
            op_desc = op["op_desc"]
            machine = op["machine"]
            parts_list = op.get("parts", "")
            is_first = (op_idx == 0)
            
            is_move = any(kw in op_desc.lower() for kw in ("move for next", "move to next", "move to dispatch", "move for dispatch"))
            
            if is_move:
                ws_out.row_dimensions[cur].height = 22.0
                sc(ws_out.cell(cur, 1), op_no, bold=True, size=12, rgb="FFFFFF", left="medium", right="thin")
                for c in range(2, 15):
                    sc(ws_out.cell(cur, c), rgb="FFFFFF", left="none", right="none")
                try: ws_out.cell(cur, 2).value = op_desc
                except AttributeError: pass
                ws_out.cell(cur, 2).font = _font(bold=True, size=12)
                ws_out.cell(cur, 2).border = _border(left="thin")
                _safe_merge(ws_out, cur, 2, cur, 14)
                cur += 1
            else:
                json_data = _match_json_key(op_desc, cp_dict)
                chars = _extract_chars(json_data) if json_data else []
                if is_first and parts_list:
                    cleaned_parts = re.sub(r'\n{2,}', '\n', parts_list).strip()
                    display_op_desc = f"{op_desc}\n{cleaned_parts}"
                else:
                    display_op_desc = op_desc
                
                if not chars:
                    ws_out.row_dimensions[cur].height = 25.0
                    sc(ws_out.cell(cur, 1), op_no, bold=True, size=12, left="medium", right="thin")
                    sc(ws_out.cell(cur, 2), display_op_desc, size=11, h="left", left="thin", right="thin")
                    sc(ws_out.cell(cur, 3), machine, size=10, h="left", left="thin", right="thin")
                    for c in range(4, 14): sc(ws_out.cell(cur, c), left="thin", right="thin")
                    sc(ws_out.cell(cur, 14), left="thin")
                    cur += 1
                else:
                    step_start = cur
                    for i, ch in enumerate(chars):
                        ws_out.row_dimensions[cur].height = 44.0
                        sc(ws_out.cell(cur, 1), op_no if i == 0 else None, bold=True, size=12, left="medium", right="thin")
                        sc(ws_out.cell(cur, 2), display_op_desc if i == 0 else None, size=11, left="thin", right="thin")
                        sc(ws_out.cell(cur, 3), machine if i == 0 else None, size=10, left="thin", right="thin")
                        sc(ws_out.cell(cur, 4), str(i + 1), size=11, left="thin", right="thin")
                        sc(ws_out.cell(cur, 5), ch.get("Product", ""), size=11, h="left", left="thin", right="thin")
                        sc(ws_out.cell(cur, 6), ch.get("Process", ""), size=11, h="left", left="thin", right="thin")
                        sc(ws_out.cell(cur, 7), ch.get("Special Characteristic Class", ""), size=11, left="thin", right="thin")
                        sc(ws_out.cell(cur, 8), ch.get("Product/Process Specification/Tolerance", ""), size=11, h="left", left="thin", right="thin")
                        sc(ws_out.cell(cur, 9), ch.get("Evaluation/Measurement Technique", ""), size=11, left="thin", right="thin")
                        sc(ws_out.cell(cur, 10), ch.get("Size", ""), size=11, left="thin", right="thin")
                        sc(ws_out.cell(cur, 11), ch.get("Freq", ""), size=11, left="thin", right="thin")
                        sc(ws_out.cell(cur, 12), ch.get("Control Method", ""), size=11, h="left", left="thin", right="thin")
                        sc(ws_out.cell(cur, 13), ch.get("Responsibility", ""), size=11, left="thin", right="thin")
                        sc(ws_out.cell(cur, 14), ch.get("Reaction Plan", ""), size=11, h="left", left="thin")
                        cur += 1
                    step_end = cur - 1
                    if step_end > step_start:
                        for col in (1, 2, 3): _safe_merge(ws_out, step_start, col, step_end, col)
                    for col in (6, 9, 10, 11, 12, 13, 14):
                        anchor = step_start
                        while anchor <= step_end:
                            anchor_val = ws_out.cell(anchor, col).value
                            end_r = anchor
                            while end_r + 1 <= step_end:
                                nv = ws_out.cell(end_r + 1, col).value
                                if (nv == anchor_val or nv is None) and anchor_val: end_r += 1
                                else: break
                            if end_r > anchor: _safe_merge(ws_out, anchor, col, end_r, col)
                            anchor = end_r + 1

    offset = cur - FOOTER_START
    for r in range(FOOTER_START, FOOTER_END + 1):
        out_r = r + offset
        ws_out.row_dimensions[out_r].height = tmpl_ws.row_dimensions[r].height or 15
        for c in range(1, 15):
            _copy_cell(tmpl_ws.cell(r, c), ws_out.cell(out_r, c))
    for mr in tmpl_ws.merged_cells.ranges:
        if FOOTER_START <= mr.min_row <= FOOTER_END:
            try: ws_out.merge_cells(start_row=mr.min_row + offset, start_column=mr.min_col, end_row=mr.max_row + offset, end_column=mr.max_col)
            except Exception: pass

# ─── SHEET 3: Q-SHEETMETAL ────────────────────────────────────────────────────
def _extract_chars(node):
    chars = []
    if isinstance(node, dict):
        if "Product" in node and isinstance(node.get("Product"), str):
            chars.append(node)
        else:
            for v in node.values():
                if isinstance(v, dict):
                    chars.extend(_extract_chars(v))
    return chars

def _match_op(op_desc, cp_dict):
    op = str(op_desc).lower().strip()
    if 'oxy' in op or 'plasma' in op:
        op = 'laser cutting'
    elif 'straightening' in op:
        op = 'rolling / strightening'
    for key in cp_dict:
        kl = key.replace('_', ' ').lower().strip()
        if kl == op or kl in op or op in kl:
            return cp_dict[key]
    return None

def _copy_header(tmpl_ws, out_ws, header_end):
    for col_letter in "ABCDEFGHIJKLMN":
        w = tmpl_ws.column_dimensions[col_letter].width
        if w:
            out_ws.column_dimensions[col_letter].width = w
    for row in tmpl_ws.iter_rows(min_row=1, max_row=header_end):
        for src in row:
            _copy_cell(src, out_ws.cell(src.row, src.column))
        out_ws.row_dimensions[row[0].row].height = tmpl_ws.row_dimensions[row[0].row].height
    for mr in tmpl_ws.merged_cells.ranges:
        if mr.max_row <= header_end:
            try:
                out_ws.merge_cells(str(mr))
            except Exception:
                pass

def process_q_sheetmetal_sheet(ws_out, tmpl_ws, feasibility_path, pfd_output_path, cp_dict, img_dir):
    """
    Build Q-Sheetmetal Parts sheet dynamically.
    """
    HEADER_END = 19
    _clear_below_header(ws_out, HEADER_END)
    # Find footer in template
    FOOTER_START = None
    for r in range(100, tmpl_ws.max_row + 1):
        if tmpl_ws.cell(r, 1).value and "Note" in str(tmpl_ws.cell(r, 1).value):
            FOOTER_START = r
            break
    if not FOOTER_START:
        FOOTER_START = 257
    FOOTER_END = tmpl_ws.max_row

    # Copy header from template
    _copy_header(tmpl_ws, ws_out, HEADER_END)
    cur_row = HEADER_END + 1

    # Load feasibility
    df_feas = pd.read_excel(feasibility_path, sheet_name=0, header=3)
    df_feas.columns = df_feas.columns.astype(str).str.strip()
    op_col_name = next((c for c in df_feas.columns if 'OPERATION' in c.upper()), None)
    if op_col_name:
        df_feas[op_col_name] = df_feas[op_col_name].ffill()
    df_feas = df_feas.dropna(subset=['Part No.', 'Part Description'], how='all')

    # Filter out BOP parts and Assemblies (Fix 3)
    # The user specifies this sheet should only contain child sheetmetal parts.
    bop_keywords = ['HARDWARE', 'BOP', 'FASTENER', 'BOUGHT OUT', 'STANDARD', 'BOUGHT-OUT']
    assy_keywords = ['ASSY', 'WA', 'W/A', 'WELDED', 'ASSEMBLY', 'PROJECT', 'PWRTAN', 'PAINTED', 'PNTD']
    
    # Filter descriptions
    df_feas = df_feas[~df_feas['Part Description'].astype(str).str.upper().apply(lambda x: any(kw in x for kw in assy_keywords))]
    # Filter commodity
    df_feas = df_feas[~df_feas['Commodity'].astype(str).str.upper().apply(lambda x: any(kw in x for kw in bop_keywords + assy_keywords))]

    # Load PFD sheetmetal operations
    df_pfd = None
    if pfd_output_path and os.path.exists(pfd_output_path):
        try:
            df_pfd = pd.read_excel(pfd_output_path, sheet_name='Q-SHEETMETAL_PARTS', header=6)
        except Exception as e:
            print(f"[CP] Warning reading PFD sheetmetal sheet: {e}")

    df_sm = df_feas # Process all filtered parts
    for part_idx, (_, frow) in enumerate(df_sm.iterrows()):
        pno  = str(frow.get('Part No.', '')).split('.')[0].strip()
        if not pno or pno.lower() == 'nan':
            continue
        desc      = str(frow.get('Part Description', ''))
        assy_op   = str(frow.get(op_col_name, '') if op_col_name else '')
        thickness = str(frow.get('Thickness/Size(mm) (As per Drawing)', ''))
        material  = str(frow.get('Material Gr. (As per Drawing)', ''))

        # Part heading row
        ws_out.row_dimensions[cur_row].height = 152.25
        sc(ws_out.cell(cur_row, 1), str(part_idx + 1), bold=True, size=14, rgb=GREEN_PART, h="center", left="medium")
        sc(ws_out.cell(cur_row, 2), f"PART NUMBER - {pno} ,DESCRIPTION-{desc}", bold=True, size=14, rgb=GREEN_PART, h="center", v="center", wrap=True)
        # Images moved from B to I (Col 9)
        for c in range(3, 9):
            sc(ws_out.cell(cur_row, c), rgb=GREEN_PART)
        _safe_merge(ws_out, cur_row, 2, cur_row, 8)
        
        for c in range(9, 13):
            sc(ws_out.cell(cur_row, c), rgb=GREEN_PART)
        _safe_merge(ws_out, cur_row, 9, cur_row, 12)
        
        # Image in Column I
        if img_dir:
            img_path = os.path.join(img_dir, f"{pno}_clean.png")
            if os.path.exists(img_path):
                try:
                    img = XLImage(img_path)
                    img.height, img.width = 135, 190
                    # Anchor to I (Col 9)
                    ws_out.add_image(img, ws_out.cell(cur_row, 9).coordinate)
                except Exception as e:
                    print(f"[CP] Image error {pno}: {e}")
        sc(ws_out.cell(cur_row, 13), rgb=GREEN_PART)
        op_info = f"This Part assemble in\n{assy_op}" if assy_op != 'nan' else ""
        sc(ws_out.cell(cur_row, 14), op_info, bold=True, size=14, rgb=GREEN_PART, h="center", v="center", wrap=True)
        cur_row += 1

        # Operation rows from PFD
        if df_pfd is not None:
            parts_col = next((c for c in df_pfd.columns if 'List of Parts' in str(c)), None)
            op_no_col = next((c for c in df_pfd.columns if str(c).strip() == 'Operation No'), None)
            op_ds_col = next((c for c in df_pfd.columns if 'Operation Description' in str(c)), None)
            mach_col  = next((c for c in df_pfd.columns if 'Machine' in str(c)), None)
            part_ops  = df_pfd[df_pfd[parts_col].astype(str).str.contains(pno, na=False)] if parts_col else pd.DataFrame()
            storage_count = 0
            for _, pfd_row in part_ops.iterrows():
                op_no   = str(pfd_row[op_no_col])  if op_no_col else ""
                op_name = str(pfd_row[op_ds_col])   if op_ds_col else ""
                machine = str(pfd_row[mach_col]).replace('nan', '') if mach_col else ""
                if 'storage' in op_name.lower():
                    storage_count += 1
                    cp_data = cp_dict.get("Storage2") if storage_count >= 2 else _match_op(op_name, cp_dict)
                else:
                    cp_data = _match_op(op_name, cp_dict)
                char_list = _extract_chars(cp_data) if cp_data else []
                if not char_list:
                    ws_out.row_dimensions[cur_row].height = 35.25
                    sc(ws_out.cell(cur_row, 1), op_no,   bold=True, size=14, h="center", left="medium")
                    sc(ws_out.cell(cur_row, 2), op_name, bold=True, size=14, h="left")
                    sc(ws_out.cell(cur_row, 3), machine, size=11,   h="left")
                    for c in range(4, 15):
                        sc(ws_out.cell(cur_row, c))
                    cur_row += 1
                else:
                    op_start = cur_row
                    last_proc = ""
                    for i, char in enumerate(char_list):
                        ws_out.row_dimensions[cur_row].height = 44.25
                        sc(ws_out.cell(cur_row, 1), op_no   if i == 0 else None, bold=True, size=14, h="center", left="medium")
                        sc(ws_out.cell(cur_row, 2), op_name if i == 0 else None, bold=True, size=14, h="left")
                        sc(ws_out.cell(cur_row, 3), machine if i == 0 else None, size=11,   h="left")
                        sc(ws_out.cell(cur_row, 4), str(i + 1), h="center")
                        prod = char.get("Product", "")
                        proc = char.get("Process", "")
                        if not proc and last_proc:
                            proc = last_proc
                        if proc:
                            last_proc = proc
                        sc(ws_out.cell(cur_row, 5), prod, h="left")
                        sc(ws_out.cell(cur_row, 6), proc, h="left")
                        sc(ws_out.cell(cur_row, 7), char.get("Special Characteristic Class", ""), h="center")
                        spec = char.get("Product/Process Specification/Tolerance", "")
                        if "Thickness" in str(prod) and not spec:
                            spec = thickness if thickness != 'nan' else ""
                        elif "Material" in str(prod) and not spec:
                            spec = material if material != 'nan' else ""
                        sc(ws_out.cell(cur_row, 8),  spec, h="left")
                        sc(ws_out.cell(cur_row, 9),  char.get("Evaluation/Measurement Technique", ""), h="left")
                        sc(ws_out.cell(cur_row, 10), char.get("Size", ""), h="center")
                        sc(ws_out.cell(cur_row, 11), char.get("Freq", ""), h="center")
                        sc(ws_out.cell(cur_row, 12), char.get("Control Method", ""), h="left")
                        sc(ws_out.cell(cur_row, 13), char.get("Responsibility", ""), h="left")
                        sc(ws_out.cell(cur_row, 14), char.get("Reaction Plan", ""), h="left")
                        cur_row += 1
                    # Vertical merges
                    if cur_row - 1 > op_start:
                        for col in range(1, 4):
                            _safe_merge(ws_out, op_start, col, cur_row - 1, col)
                        for col in [6] + list(range(8, 15)):
                            ms = op_start
                            while ms <= cur_row - 1:
                                val = ws_out.cell(ms, col).value
                                me  = ms
                                while me + 1 <= cur_row - 1 and ws_out.cell(me + 1, col).value == val and val and str(val).strip():
                                    me += 1
                                if me > ms:
                                    _safe_merge(ws_out, ms, col, me, col)
                                ms = me + 1

        # Move to next process row
        ws_out.row_dimensions[cur_row].height = 39.75
        sc(ws_out.cell(cur_row, 1), "120", bold=True, size=14, h="center", left="medium")
        sc(ws_out.cell(cur_row, 2), "Move to the next process", bold=True, size=14, h="center")
        for c in range(3, 15):
            sc(ws_out.cell(cur_row, c))
        _safe_merge(ws_out, cur_row, 2, cur_row, 14)
        cur_row += 1

    # Copy footer from template
    footer_offset = cur_row - FOOTER_START
    for r in range(FOOTER_START, FOOTER_END + 1):
        out_r = r + footer_offset
        ws_out.row_dimensions[out_r].height = tmpl_ws.row_dimensions[r].height or 18
        for c in range(1, 15):
            _copy_cell(tmpl_ws.cell(r, c), ws_out.cell(out_r, c))
    for mr in tmpl_ws.merged_cells.ranges:
        if FOOTER_START <= mr.min_row <= FOOTER_END:
            _safe_merge(ws_out, mr.min_row + footer_offset, mr.min_col, mr.max_row + footer_offset, mr.max_col)

# ─── SHEET 4: T-BOP & HARDWARES ───────────────────────────────────────────────
def _find_t_marker(ws, default=19):
    for r in range(1, 60):
        if str(ws.cell(r, 1).value).strip() == 'T':
            return r
    return default

def _capture_bop_footer(ws, footer_start_row):
    footer_data, footer_merges = [], []
    for r in range(footer_start_row, ws.max_row + 1):
        row_cells = []
        for c in range(1, 15):
            cell = ws.cell(r, c)
            row_cells.append({'val': cell.value, 'style': cell._style, 'height': ws.row_dimensions[r].height})
        footer_data.append(row_cells)
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= footer_start_row:
            footer_merges.append({'min_col': m.min_col, 'max_col': m.max_col,
                                   'min_row_off': m.min_row - footer_start_row,
                                   'max_row_off': m.max_row - footer_start_row})
            ws.unmerge_cells(str(m))
    return footer_data, footer_merges

def _redraw_bop_footer(ws, cur_row, footer_data, footer_merges):
    for i, row_data in enumerate(footer_data):
        tgt_r = cur_row + i
        ws.row_dimensions[tgt_r].height = row_data[0]['height']
        for j, ci in enumerate(row_data):
            cell = ws.cell(tgt_r, j + 1)
            cell.value   = ci['val']
            cell._style  = ci['style']
    for m in footer_merges:
        _safe_merge(ws, cur_row + m['min_row_off'], m['min_col'], cur_row + m['max_row_off'], m['max_col'])

def process_bop_sheet(ws, bop_data):
    """
    Inject BOP control plan data into T - BOP & Hardwares sheet.
    bop_data: dict with 'processes' list (matches cp_bop_dict.json structure).
    """
    processes = bop_data.get("processes", [])
    t_marker_row = _find_t_marker(ws)
    data_start   = t_marker_row + 1

    # Find footer
    footer_start = None
    for r in range(data_start, ws.max_row + 1):
        v = str(ws.cell(r, 1).value).upper()
        if "NOTE" in v or "PREPARED BY" in v:
            footer_start = r
            break
    if not footer_start:
        footer_start = ws.max_row + 1

    footer_data, footer_merges = _capture_bop_footer(ws, footer_start)
    ws.delete_rows(data_start, ws.max_row - data_start + 1)

    std = _std_border()
    cur_row = data_start

    for proc in processes:
        part_no   = proc.get("part_process_number")
        desc      = proc.get("process_name_operation_description", "")
        machine   = proc.get("machine_device_jig_tools_for_mfg", "")
        chars     = proc.get("characteristics", [])
        has_items = "List of items" in str(desc)
        block_start = cur_row

        if not chars:
            ws.row_dimensions[cur_row].height = 30
            ws.cell(cur_row, 1).value = part_no
            ws.cell(cur_row, 2).value = desc
            _safe_merge(ws, cur_row, 2, cur_row, 14)
            for c in range(1, 15):
                ws.cell(cur_row, c).border    = std
                ws.cell(cur_row, c).alignment = _align()
                ws.cell(cur_row, c).font      = _font()
            cur_row += 1
            continue

        for i, char in enumerate(chars):
            ws.row_dimensions[cur_row].height = 110 if has_items else 45
            if i == 0:
                ws.cell(cur_row, 1).value = part_no
                ws.cell(cur_row, 2).value = desc
                ws.cell(cur_row, 3).value = machine
            ws.cell(cur_row, 4).value  = char.get("no")
            ws.cell(cur_row, 5).value  = char.get("product")
            ws.cell(cur_row, 6).value  = char.get("process")
            ws.cell(cur_row, 7).value  = char.get("special_characteristic_class")
            ws.cell(cur_row, 8).value  = char.get("product_process_specification_tolerance")
            ws.cell(cur_row, 9).value  = char.get("evaluation_measurement_technique")
            ws.cell(cur_row, 10).value = char.get("sample_size")
            ws.cell(cur_row, 11).value = char.get("sample_freq")
            ws.cell(cur_row, 12).value = char.get("control_method")
            ws.cell(cur_row, 13).value = char.get("responsibility")
            ws.cell(cur_row, 14).value = char.get("reaction_plan")
            for c in range(1, 15):
                ws.cell(cur_row, c).border    = std
                ws.cell(cur_row, c).alignment = _align(wrap=True)
                ws.cell(cur_row, c).font      = _font()
            cur_row += 1

        if len(chars) > 1:
            _safe_merge(ws, block_start, 1, cur_row - 1, 1)
            _safe_merge(ws, block_start, 2, cur_row - 1, 2)
            _safe_merge(ws, block_start, 3, cur_row - 1, 3)

    _redraw_bop_footer(ws, cur_row, footer_data, footer_merges)

def transfuse_sheet_structure(tmpl_ws, target_ws, header_rows=0):
    """
    Clones column widths and header rows from template to a new worksheet.
    """
    # Clone column dimensions
    for c_idx in range(1, 40): # Buffer range
        col_letter = openpyxl.utils.get_column_letter(c_idx)
        width = tmpl_ws.column_dimensions[col_letter].width
        if width:
            target_ws.column_dimensions[col_letter].width = width
        else:
            # Default fallback for structure columns if not set
            if c_idx < 15: target_ws.column_dimensions[col_letter].width = 12

    # Clone header rows exactly
    if header_rows > 0:
        for r in range(1, header_rows + 1):
            target_ws.row_dimensions[r].height = tmpl_ws.row_dimensions[r].height
            for c in range(1, 37):
                _copy_cell(tmpl_ws.cell(r, c), target_ws.cell(r, c))
        
        # Re-apply merges strictly for the header block
        for mr in tmpl_ws.merged_cells.ranges:
            if mr.max_row <= header_rows:
                try:
                    target_ws.merge_cells(str(mr))
                except: pass
    
    # Clone images that are anchored within the header area
    if hasattr(tmpl_ws, "_images"):
        for img in tmpl_ws._images:
            try:
                # Basic anchor detection
                row_idx = img.anchor._from.row if hasattr(img.anchor, '_from') else 0
                if row_idx < header_rows:
                    from copy import copy
                    new_img = copy(img)
                    target_ws.add_image(new_img)
            except: pass

# ─── MAIN ORCHESTRATOR ────────────────────────────────────────────────────────
def generate_control_plan_excel(
    feasibility_path: str,
    template_path: str,
    output_path: str,
    dict_dir: str,
    pfd_output_path: str = "",
    img_dir: str = "",
    assy_data: list = None,
    pfd_img_dir: str = "",
):
    """
    Generate full Control Plan Excel workbook with all 4 sheets using the Fresh Workbook approach.
    """
    tmpl_wb = openpyxl.load_workbook(template_path)
    new_wb = openpyxl.Workbook()
    for sheet in new_wb.sheetnames:
        del new_wb[sheet] # Clear default sheets

    # ── Sheet 1: Sub Assmbly Index ─────────────────────────────────────────────
    if SUB_ASSY_SHEET in tmpl_wb.sheetnames:
        ws_out = new_wb.create_sheet(SUB_ASSY_SHEET)
        transfuse_sheet_structure(tmpl_wb[SUB_ASSY_SHEET], ws_out, header_rows=2)
        if assy_data:
            print("[CP] Processing Sub Assmbly Index sheet...")
            idx_img_dir = pfd_img_dir if pfd_img_dir else img_dir
            process_sub_assy_index_sheet(ws_out, assy_data, img_dir=idx_img_dir)

    # ── Sheet 2: ASSY_SUB_ASSY ────────────────────────────────────────────────
    if ASSY_ASSY_SHEET in tmpl_wb.sheetnames:
        ws_out = new_wb.create_sheet(ASSY_ASSY_SHEET)
        transfuse_sheet_structure(tmpl_wb[ASSY_ASSY_SHEET], ws_out, header_rows=7)
        if pfd_output_path and os.path.exists(pfd_output_path):
            assy_dict_path = os.path.join(dict_dir, "assy", "cp_assy_dict.json")
            if os.path.exists(assy_dict_path):
                with open(assy_dict_path, encoding="utf-8") as f:
                    assy_dict = json.load(f)
                print("[CP] Processing ASSY_SUB_ASSY sheet...")
                process_assy_sub_assy_sheet(ws_out, tmpl_wb[ASSY_ASSY_SHEET], pfd_output_path, assy_dict)

    # ── Sheet 3: Q-Sheetmetal Parts ───────────────────────────────────────────
    if Q_SM_SHEET in tmpl_wb.sheetnames:
        ws_out = new_wb.create_sheet(Q_SM_SHEET)
        # Note: process_q_sheetmetal_sheet handles its own combined structural copy/injection
        sm_dict_path = os.path.join(dict_dir, "sheetmetal", "cp_sheetmetal_dict.json")
        if os.path.exists(sm_dict_path):
            with open(sm_dict_path, encoding="utf-8") as f:
                sm_dict = json.load(f)
            print("[CP] Processing Q-Sheetmetal Parts sheet...")
            process_q_sheetmetal_sheet(ws_out, tmpl_wb[Q_SM_SHEET], feasibility_path, pfd_output_path, sm_dict, img_dir)

    # ── Sheet 4: T-BOP & Hardwares ────────────────────────────────────────────
    if BOP_SHEET in tmpl_wb.sheetnames:
        ws_out = new_wb.create_sheet(BOP_SHEET)
        transfuse_sheet_structure(tmpl_wb[BOP_SHEET], ws_out, header_rows=6)
        bop_dict_path = os.path.join(dict_dir, "bop", "cp_bop_dict.json")
        if os.path.exists(bop_dict_path):
            with open(bop_dict_path, encoding="utf-8") as f:
                bop_data = json.load(f)
            print("[CP] Processing T-BOP & Hardwares sheet...")
            process_bop_sheet(ws_out, bop_data)

    new_wb.save(output_path)
    print(f"[CP] Clean Control Plan saved -> {output_path}")
    return output_path
