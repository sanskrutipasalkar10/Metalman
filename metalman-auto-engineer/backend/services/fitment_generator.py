"""
fitment_generator.py
Metalman Auto Engineer — Fitment Checksheet Generation Service

Generates the Part Fitment Check Sheet Excel document from:
  - An NPD Feasibility Matrix (.xlsx)
  - A Fitment Checksheet Excel template
  - Part images extracted from the drawings/upload directory

USAGE NOTE:
  Place the blank Fitment Checksheet template at:
  backend/assets/fitment_checksheet_template.xlsx
"""

import os
import shutil
import re
import datetime
from copy import copy

import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, XDRPositiveSize2D
from openpyxl.styles import Alignment


# ==========================================
# UTILITIES: PLACEMENT & COORDINATES
# ==========================================

def _px_to_emu(pixels: int) -> int:
    """Converts pixels to English Metric Units for precise Excel image positioning."""
    return int(pixels * 9525)


def _update_header_safe(ws, cell_coord: str, value) -> None:
    """
    Updates a cell value safely, even if the cell is part of a merged range.
    Writes to the top-left (master) cell of the merged region.
    """
    cell = ws[cell_coord]
    if type(cell).__name__ == "MergedCell":
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in merged_range:
                ws[merged_range.coord.split(":")[0]].value = value
                break
    else:
        ws[cell_coord].value = value


# ==========================================
# DATA HARVESTER
# ==========================================

def _extract_feasibility_data(feasibility_xlsx: str):
    """
    Extracts BOM-level rows from the NPD Feasibility Matrix Excel file.

    Returns:
        bom_data   (list[dict]):  list of part rows with s_no, part_no, part_name, qty
        main_no    (str):         master part number (extracted from filename, then file content)
        main_desc  (str):         description of first row (main assembly description)
    """
    try:
        df = pd.read_excel(feasibility_xlsx, skiprows=3, engine="openpyxl")
        df.columns = df.columns.astype(str).str.strip().str.replace("\n", "")

        # Extract master part number — priority: 8-9 digit number in filename
        filename_match = re.search(r"\d{8,9}", os.path.basename(feasibility_xlsx))
        if filename_match:
            main_part_no = filename_match.group(0)
        else:
            # Fallback: any digit sequence in filename
            fallback = re.search(r"\d+", os.path.basename(feasibility_xlsx))
            main_part_no = fallback.group(0) if fallback else "MASTER"

        bom_data = []
        for _, row in df.iterrows():
            if pd.isna(row.get("Sr No")):
                continue
            p_no = str(row.get("Part No.")).split(".")[0].replace(".0", "").strip()
            if p_no.lower() == "nan" or p_no == "":
                continue

            bom_data.append({
                "s_no":      row.get("Sr No"),
                "part_no":   p_no,
                "part_name": row.get("Part Description"),
                "qty":       row.get("Qty/Assy"),
            })

        main_desc = str(df.iloc[0].get("Part Description", "ASSEMBLY")) if len(df) > 0 else "ASSEMBLY"
        return bom_data, main_part_no, main_desc

    except Exception as e:
        print(f"[fitment_generator] ERROR extracting feasibility data: {e}")
        return [], "MASTER", "ASSEMBLY"


# ==========================================
# CHECK SHEET GENERATOR ENGINE
# ==========================================

def generate_fitment_check_sheet(
    feasibility_file: str,
    template_path: str,
    output_path: str,
    image_dir: str,
) -> bool:
    """
    Generates the Fitment Checksheet Excel from the feasibility data and part images.

    Args:
        feasibility_file:  Path to the NPD Feasibility Matrix .xlsx
        template_path:     Path to the blank Fitment Checksheet template .xlsx
        output_path:       Destination path for the generated output .xlsx
        image_dir:         Directory containing part images named {part_no}_clean.png
                           or {part_no}_raw.png (falls back gracefully if not found)

    Returns:
        True on success, False on failure.
    """
    # ── Validate template exists ──────────────────────────────────────────────
    if not os.path.exists(template_path):
        print(f"[fitment_generator] WARNING: Template not found at '{template_path}'. Skipping generation.")
        return False

    # ── Extract data ──────────────────────────────────────────────────────────
    bom_data, main_no, main_desc = _extract_feasibility_data(feasibility_file)
    if not bom_data:
        print("[fitment_generator] WARNING: No BOM rows extracted. Skipping generation.")
        return False

    print(f"[fitment_generator] Generating Fitment Checksheet for part: {main_no} ({len(bom_data)} rows)")

    try:
        # ── Copy template to output path ─────────────────────────────────────
        shutil.copy(template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # ── A. Fill Header ───────────────────────────────────────────────────
        today = datetime.datetime.now().strftime("%d-%m-%Y")
        _update_header_safe(ws, "C3", "Final Assy")
        _update_header_safe(ws, "C4", "952-NX")
        _update_header_safe(ws, "C5", main_desc)
        _update_header_safe(ws, "C6", main_no)
        _update_header_safe(ws, "J4", today)

        # ── B. Data injection setup ──────────────────────────────────────────
        start_data_row = 9
        current_row = start_data_row

        # Column D (index 3) visual dimensions for image centering
        col_width_px  = 165          # approximate column D width in pixels
        row_height_pts = 75          # row height in points
        row_height_px  = row_height_pts * 1.33  # convert to pixels

        # Capture merge patterns from template row 9 to replicate on inserted rows
        anchor_merges = [
            (m.min_col, m.max_col)
            for m in ws.merged_cells.ranges
            if m.min_row == 9
        ]
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── C. Row-by-row data + image injection ─────────────────────────────
        for i, item in enumerate(bom_data):
            # Insert a new row for every item after the first
            if i > 0:
                ws.insert_rows(current_row)
                # Re-apply merge patterns from the template row
                for min_c, max_c in anchor_merges:
                    ws.merge_cells(
                        start_row=current_row, end_row=current_row,
                        start_column=min_c, end_column=max_c,
                    )
                # Copy cell styles (borders / fonts) from the row above
                for c in range(1, 12):
                    ws.cell(row=current_row, column=c)._style = copy(
                        ws.cell(row=current_row - 1, column=c)._style
                    )

            # Write text data columns
            ws.cell(row=current_row, column=1).value = item["s_no"]
            ws.cell(row=current_row, column=2).value = item["part_name"]
            ws.cell(row=current_row, column=3).value = item["part_no"]
            ws.cell(row=current_row, column=5).value = item["qty"]

            # Clear any template placeholder text in image cell and remark columns
            ws.cell(row=current_row, column=4).value = ""
            for c in [6, 7, 8, 9, 10, 11]:
                ws.cell(row=current_row, column=c).value = ""

            # Apply centred alignment to text columns
            for c_idx in [1, 2, 3, 5]:
                ws.cell(row=current_row, column=c_idx).alignment = center_align

            # ── D. IMAGE INSERTION (Column D = Col Index 3) ──────────────────
            p_no = item["part_no"]
            img_path = os.path.join(image_dir, f"{p_no}_clean.png")
            if not os.path.exists(img_path):
                img_path = os.path.join(image_dir, f"{p_no}_raw.png")

            if os.path.exists(img_path):
                print(f"   [fitment] Found image for {p_no}: {img_path}")
                img = ExcelImage(img_path)

                # Target visual size (px)
                target_w, target_h = 130, 90
                img.width  = target_w
                img.height = target_h
                ws.row_dimensions[current_row].height = 85

                # Use direct cell coordinate for more reliable insertion
                photo_cell = f"D{current_row}"
                ws.add_image(img, photo_cell)
                print(f"   [fitment] Successfully injected {photo_cell}")
            else:
                print(f"   [fitment] WARNING: No image found for {p_no} at {img_path}")
                # No image — use a compact row height
                ws.row_dimensions[current_row].height = 35

            current_row += 1

        # ── E. Snapping footer logic ─────────────────────────────────────────
        # The footer (REMARK row) was pushed down by insert_rows calls.
        # Find it and delete any gap rows between the last data row and the footer.
        new_footer_row = 0
        for row in range(current_row, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and "REMARK" in str(cell_val).upper():
                new_footer_row = row
                break

        if new_footer_row > current_row:
            rows_to_delete = new_footer_row - current_row
            ws.delete_rows(current_row, rows_to_delete)

        wb.save(output_path)
        print(f"[fitment_generator] COMPLETE → {output_path}")
        return True

    except Exception as e:
        import traceback
        print(f"[fitment_generator] ERROR during generation: {e}")
        traceback.print_exc()
        return False
