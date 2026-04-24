import os
import shutil
import fitz  # PyMuPDF
from PIL import Image, ImageDraw, ImageFont
import string
import re
import io
import base64
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import cv2
import numpy as np
import datetime
from openai import OpenAI

# Initialize the OpenAI client pointing to the OLLAMA_BASE_URL.
OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434/v1")
OLLAMA_API_KEY = "ollama"  
VISION_MODEL = "qwen3-vl:235b-instruct-cloud" 

client = OpenAI(
    base_url=OLLAMA_BASE_URL,
    api_key=OLLAMA_API_KEY
)

def encode_image(pil_img):
    buffered = io.BytesIO()
    pil_img.save(buffered, format="PNG")
    return base64.b64encode(buffered.getvalue()).decode('utf-8')

# ==========================================
# OPENCV SCRIPT (SURGICAL CLEANUP)
# ==========================================
def isolate_part_and_remove_text(image_path, output_filename, target_cx, target_cy):
    print(f"   OpenCV: Engaging Surgical Cleanup for {image_path}...")
    
    img = cv2.imread(image_path)
    if img is None: return False
        
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape
    
    _, thresh = cv2.threshold(gray, 220, 255, cv2.THRESH_BINARY_INV)
    
    margin = 10
    thresh[0:margin, :] = 0
    thresh[h-margin:h, :] = 0
    thresh[:, 0:margin] = 0
    thresh[:, w-margin:w] = 0
    
    contours_raw, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    clean_mask = np.zeros_like(thresh)
    
    for cnt in contours_raw:
        x, y, w_box, h_box = cv2.boundingRect(cnt)
        if w_box > 45 or h_box > 45:
            cv2.drawContours(clean_mask, [cnt], -1, 255, thickness=cv2.FILLED)

    kernel = np.ones((12, 12), np.uint8) 
    dilated = cv2.dilate(clean_mask, kernel, iterations=2)
    
    final_contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    best_box = None
    min_distance = float('inf')
    
    for cnt in final_contours:
        x, y, w_box, h_box = cv2.boundingRect(cnt)
        if x <= 5 or y <= 5 or (x + w_box) >= (w - 5) or (y + h_box) >= (h - 5): continue
            
        obj_cx = x + (w_box // 2)
        obj_cy = y + (h_box // 2)
        distance = ((obj_cx - target_cx) ** 2 + (obj_cy - target_cy) ** 2) ** 0.5
        
        if distance < min_distance:
            min_distance = distance
            best_box = (x, y, w_box, h_box)

    if best_box is None and len(final_contours) > 0:
        for cnt in final_contours:
            x, y, w_box, h_box = cv2.boundingRect(cnt)
            obj_cx = x + (w_box // 2)
            obj_cy = y + (h_box // 2)
            distance = ((obj_cx - target_cx) ** 2 + (obj_cy - target_cy) ** 2) ** 0.5
            if distance < min_distance:
                min_distance = distance
                best_box = (x, y, w_box, h_box)

    if best_box:
        x, y, w_box, h_box = best_box
        pad = 15 
        x_start, y_start = max(0, x - pad), max(0, y - pad)
        x_end, y_end = min(w, x + w_box + pad), min(h, y + h_box + pad)
        final_cropped = img[y_start:y_end, x_start:x_end]
        cv2.imwrite(output_filename, final_cropped)
        print(f"   SUCCESS: Clean surgical extraction saved -> {output_filename}")
        return True
    else:
        print("   FATAL: OpenCV found nothing after cleanup.")
        return False

# ==========================================
# VISION AGENT SCRIPT: OLLAMA QWEN3-VL
# ==========================================
def extract_isometric_view(pdf_path, output_image_path, grid_size=10):
    print(f"   Processing PDF: {pdf_path}")
    try:
        doc = fitz.open(pdf_path)
        page = doc.load_page(0)
        pix = page.get_pixmap(dpi=200) 
        clean_img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")
        
        gridded_img = clean_img.copy()
        draw = ImageDraw.Draw(gridded_img, 'RGBA')
        w, h = gridded_img.size
        cell_w, cell_h = w / grid_size, h / grid_size
        
        for i in range(1, grid_size):
            draw.line([(i * cell_w, 0), (i * cell_w, h)], fill=(0, 200, 0, 200), width=5)
            draw.line([(0, i * cell_h), (w, i * cell_h)], fill=(0, 200, 0, 200), width=5)

        cols = list(string.ascii_uppercase)[:grid_size]
        try: font = ImageFont.truetype("arial.ttf", 65)
        except IOError: font = ImageFont.load_default()

        for row in range(grid_size):
            for col in range(grid_size):
                text_x = (col * cell_w) + (cell_w / 2) - 30
                text_y = (row * cell_h) + (cell_h / 2) - 30
                grid_label = f"{cols[col]}{row + 1}"
                draw.text((text_x, text_y), grid_label, fill=(0, 0, 0, 255), 
                          font=font, stroke_width=4, stroke_fill=(255, 255, 255, 255))

        print(f"   Asking {VISION_MODEL} (via Ollama) for coordinates...")
        base64_image = encode_image(gridded_img)
        
        prompt = """
        You are a highly precise machine vision system. Analyze this engineering CAD drawing.
        Find the bounding box that fully encloses exactly ONE 3D isometric (pictorial) view.
        
        CRITICAL RULES:
        1. The 3D isometric view shows depth, angles, and 3D volume.
        2. ANTI-DIMENSION RULE: The 3D isometric view is for visual reference only. It NEVER has dimension lines, arrows, or measurement numbers on it.
        3. STRICTLY IGNORE flat 2D orthographic views (top, front, side views). These are usually covered in dimensions and text.
        4. The isometric view is usually floating in an open space, often in the top-right or bottom-right area.
        
        Output ONLY the top-left grid cell and the bottom-right grid cell separated by a hyphen.
        Do not include any explanation. Example: G2-J5
        """
        
        print(f"   [VISION] Calling {VISION_MODEL} for isometric extraction...", flush=True)
        import time
        start_t = time.time()
        
        response = client.chat.completions.create(
            model=VISION_MODEL,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_image}"}}
                    ]
                }
            ],
            temperature=0.0,  
            top_p=0.1
        )
        
        elapsed = time.time() - start_t
        print(f"   [VISION] AI responded in {elapsed:.2f}s.", flush=True)
        
        ai_response_text = response.choices[0].message.content.strip()
        print(f"   [VISION] Raw Response: {ai_response_text}", flush=True)
        match = re.search(r'([A-Z])(\d+)-([A-Z])(\d+)', ai_response_text)
        
        if not match: return None
            
        col_1, row_1, col_2, row_2 = match.groups()
        
        col_start_idx = min(string.ascii_uppercase.index(col_1), string.ascii_uppercase.index(col_2))
        col_end_idx = max(string.ascii_uppercase.index(col_1), string.ascii_uppercase.index(col_2))
        row_start_idx = min(int(row_1) - 1, int(row_2) - 1)
        row_end_idx = max(int(row_1) - 1, int(row_2) - 1)
        
        ai_center_x = ((col_start_idx + col_end_idx + 1) / 2) * cell_w
        ai_center_y = ((row_start_idx + row_end_idx + 1) / 2) * cell_h
        
        BUFFER = 1
        crop_left_idx = max(0, col_start_idx - BUFFER)
        crop_right_idx = min(grid_size - 1, col_end_idx + BUFFER)
        crop_top_idx = max(0, row_start_idx - BUFFER)
        crop_bottom_idx = min(grid_size - 1, row_end_idx + BUFFER)
        
        left, top = crop_left_idx * cell_w, crop_top_idx * cell_h
        right, bottom = (crop_right_idx + 1) * cell_w, (crop_bottom_idx + 1) * cell_h
        
        target_cx = int(ai_center_x - left)
        target_cy = int(ai_center_y - top)
        
        clean_img.crop((left, top, right, bottom)).save(output_image_path)
        return (target_cx, target_cy)

    except Exception as e:
        print(f"   Vision Error: {e}")
        return None

# ==========================================
# DATA HARVESTER: EXCEL TO PYTHON
# ==========================================
def extract_feasibility_data(feasibility_xlsx):
    print(f"\nScanning Feasibility Document: {feasibility_xlsx}...", flush=True)
    try:
        # === STEP 1: Extract Master Part Number (Priority: Filename 8-9 digits) ===
        match = re.search(r'\d{8,9}', os.path.basename(feasibility_xlsx))
        if not match:
            match = re.search(r'\d+', os.path.basename(feasibility_xlsx))
        main_part_no = match.group(0) if match else "MASTER"
        main_part_desc = "ASSEMBLY"

        # === STEP 2: Read BOM data rows ===
        df = pd.read_excel(feasibility_xlsx, skiprows=3, engine='openpyxl')
        df.columns = df.columns.astype(str).str.strip().str.replace('\n', '')

        # === STEP 3: Find Level 0 description ===
        level_cols  = [c for c in df.columns if 'BOM LEVEL' in str(c).upper() or (str(c).upper().strip() == 'LEVEL')]
        desc_cols   = [c for c in df.columns if 'DESCRIPTION' in str(c).upper() or 'PART DESC' in str(c).upper()]
        part_no_cols= [c for c in df.columns if 'PART NO' in str(c).upper()]

        found_desc = False
        if level_cols and desc_cols:
            level_col = level_cols[0]
            desc_col  = desc_cols[0]
            # Try to find BOM Level == 0 row explicitly
            level0_rows = df[df[level_col].astype(str).str.strip().isin(['0', '0.0'])]
            if not level0_rows.empty:
                main_part_desc = str(level0_rows.iloc[0][desc_col]).strip()
                # Also grab part no from Level 0 if filename had no match
                if main_part_no == "MASTER" and part_no_cols:
                    lvl0_pno = str(level0_rows.iloc[0][part_no_cols[0]]).strip().replace('.0', '')
                    if lvl0_pno and lvl0_pno.lower() not in ['nan', '']:
                        main_part_no = lvl0_pno
                found_desc = True

        if not found_desc and part_no_cols and desc_cols:
            # Search data rows for the main part number
            main_row = df[df[part_no_cols[0]].astype(str).str.contains(main_part_no, na=False)]
            if not main_row.empty:
                main_part_desc = str(main_row.iloc[0][desc_cols[0]]).strip()
                found_desc = True

        if not found_desc:
            # Last resort: scan the pre-header rows (rows 1-3) of the raw Excel
            try:
                df_hdr = pd.read_excel(feasibility_xlsx, nrows=3, header=None, engine='openpyxl')
                for _, hrow in df_hdr.iterrows():
                    for val in hrow:
                        if isinstance(val, str) and len(val) > 4 and val.strip().lower() not in ['nan', '']:
                            candidate = val.strip()
                            # Prefer cells that look like a product description (not all digits/symbols)
                            if not candidate.replace(' ', '').isdigit():
                                main_part_desc = candidate
                                found_desc = True
                                break
                    if found_desc:
                        break
            except Exception:
                pass

        print(f"Extracted Global Part No: {main_part_no} | Desc: {main_part_desc}", flush=True)

        # === STEP 4: Build BOM rows (sub-parts only) ===
        bom_data = []
        for index, row in df.iterrows():
            if pd.isna(row.get('Sr No')) or str(row.get('Sr No')).strip() == '': continue
            part_number = str(row.get('Part No.')).strip()
            if part_number.lower() == 'nan' or part_number == '': continue
            if part_number.endswith('.0'): part_number = part_number[:-2]

            bom_data.append({
                "s_no":    row.get('Sr No'),
                "level":   row.get('BOM Level'),
                "part_no": part_number,
                "part_name": row.get('Part Description'),
                "rev":     row.get('Revision no'),
                "qty":     row.get('Qty/Assy'),
                "remarks": row.get('Commodity'),
                "thk":     row.get('Thickness/Size(mm) (As per Drawing)'),
                "grade":   row.get('Material Gr. (As per Drawing)'),
                "process": row.get('Manufacturing Process Details')
            })

        print(f"Extracted {len(bom_data)} valid sub-parts from Feasibility Doc.", flush=True)
        return bom_data, main_part_no, main_part_desc
    except Exception as e:
        print(f"Harvester Error: {e}", flush=True)
        return [], "MASTER", "ASSEMBLY"


# ==========================================
# INJECTOR: MASTER EXCEL WRITER
# ==========================================
def update_header_safe(sheet, cell_coord, value):
    """Safely updates a header cell WITHOUT breaking its merge formatting."""
    cell = sheet[cell_coord]
    if type(cell).__name__ == 'MergedCell':
        for merged_range in sheet.merged_cells.ranges:
            if cell_coord in merged_range:
                top_left = merged_range.coord.split(':')[0]
                sheet[top_left].value = value
                break
    else:
        cell.value = value

def safe_write_data(sheet, cell_coord, value):
    """Used strictly for the data table to smash through the footer if needed."""
    cell = sheet[cell_coord]
    if type(cell).__name__ == 'MergedCell':
        for merged_range in list(sheet.merged_cells.ranges):
            if cell_coord in merged_range:
                sheet.unmerge_cells(str(merged_range))
                sheet[cell_coord].value = value
                break
    else:
        cell.value = value


def generate_bom_excel(feasibility_file: str, template_path: str, pdf_directory: str, tmp_image_directory: str, output_path: str, force_rewrite: bool = True):
    """
    Main orchestrator for generating the BOM Excel internally, using the extracted data
    and CAD PDFs. Dynamic paths allow execution via FastAPI endpoints.
    """
    bom_data, main_part_no, main_part_desc = extract_feasibility_data(feasibility_file)
    if not bom_data:
        return {"status": "error", "message": "Failed to extract feasibility data."}

    print(f"\nBooting Master BOM Injector...", flush=True)

    # ✅ PRESERVE TEMPLATE: Copy to output first, then modify the copy.
    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    sheet = wb.active

    today_date = datetime.datetime.now().strftime("%d/%m/%Y")
    print(f"   Generating Header Data: Date={today_date}, Part No={main_part_no}", flush=True)

    # 1. Update Top-Right Corner Date
    found_rev_date = False
    for row in range(1, 4):
        for col in range(8, 13):  # Scan columns H through L
            cell_val = sheet.cell(row=row, column=col).value
            if cell_val and isinstance(cell_val, str) and "Rev Date:" in cell_val:
                if today_date not in cell_val:
                    sheet.cell(row=row, column=col).value = cell_val.replace("Rev Date:", f"Rev Date: {today_date}")
                found_rev_date = True
                break
        if found_rev_date:
            break

    if not found_rev_date:
        update_header_safe(sheet, 'J1', f"Format No - F/NPD/05\nRev No : 00\nRev Date: {today_date}")

    # 2. Insert Part Number & Description into the green header band (Row 4).
    # 🛡️ FIX: Don't hardcode 'A4' — scan ALL merged ranges in row 4 and pick the
    # widest one (largest column span). This correctly targets the centre banner cell
    # regardless of exact template layout.
    header_str = f"{main_part_no} - {main_part_desc}"
    green_cell_coord = 'A4'  # safe default
    best_span = -1
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row == 4 and merged_range.max_row == 4:
            span = merged_range.max_col - merged_range.min_col
            if span > best_span:
                best_span = span
                # top-left cell of the merged range
                green_cell_coord = f"{openpyxl.utils.get_column_letter(merged_range.min_col)}{merged_range.min_row}"
    print(f"   Writing '{header_str}' -> green cell {green_cell_coord}", flush=True)
    update_header_safe(sheet, green_cell_coord, header_str)

    current_row = 6 
    
    for item in bom_data:
        part_no = item.get('part_no')
        print(f"\nProcessing Part: {part_no}")
        
        # Use safe_write_data for the actual BOM table rows
        safe_write_data(sheet, f'A{current_row}', item.get('s_no', ''))
        safe_write_data(sheet, f'B{current_row}', item.get('level', ''))
        safe_write_data(sheet, f'C{current_row}', part_no)
        safe_write_data(sheet, f'D{current_row}', "NA")
        safe_write_data(sheet, f'E{current_row}', item.get('part_name', ''))
        safe_write_data(sheet, f'G{current_row}', item.get('rev', ''))
        safe_write_data(sheet, f'H{current_row}', item.get('qty', ''))
        safe_write_data(sheet, f'I{current_row}', item.get('remarks', ''))
        safe_write_data(sheet, f'J{current_row}', item.get('thk', ''))
        safe_write_data(sheet, f'K{current_row}', item.get('grade', ''))
        safe_write_data(sheet, f'L{current_row}', item.get('process', ''))
        
        pdf_target = None
        if os.path.exists(pdf_directory):
            for file in os.listdir(pdf_directory):
                if part_no in file and file.lower().endswith(".pdf"):
                    pdf_target = os.path.join(pdf_directory, file)
                    break
                
        raw_image_path = os.path.join(tmp_image_directory, f"{part_no}_raw.png")
        clean_image_path = os.path.join(tmp_image_directory, f"{part_no}_clean.png")
        
        if pdf_target and (force_rewrite or not os.path.exists(clean_image_path)):
            print(f"   Executing Qwen3-VL extraction for {part_no}...")
            ai_target = extract_isometric_view(pdf_target, raw_image_path)
            
            if ai_target and os.path.exists(raw_image_path):
                target_cx, target_cy = ai_target
                isolate_part_and_remove_text(raw_image_path, clean_image_path, target_cx, target_cy)
                
        target_image = clean_image_path if os.path.exists(clean_image_path) else raw_image_path
        
        if os.path.exists(target_image):
            print(f"   Injecting Photo into row {current_row}...")
            img = ExcelImage(target_image)
            img.width, img.height = 140, 100 
            sheet.row_dimensions[current_row].height = 80
            sheet.column_dimensions['F'].width = 22
            
            photo_cell = f'F{current_row}'
            if type(sheet[photo_cell]).__name__ == 'MergedCell':
                for merged_range in list(sheet.merged_cells.ranges):
                    if photo_cell in merged_range:
                        sheet.unmerge_cells(str(merged_range))
                        break
            
            sheet.add_image(img, photo_cell)
        else:
            print(f"   No PDF found. Skipping Photo for {part_no}.")
            sheet.row_dimensions[current_row].height = 25 
            
        current_row += 1

    wb.save(output_path)
    print(f"\nPIPELINE COMPLETE: Final BOM saved as -> {output_path}")
    return {"status": "success", "bom_path": output_path, "part_no": main_part_no}
