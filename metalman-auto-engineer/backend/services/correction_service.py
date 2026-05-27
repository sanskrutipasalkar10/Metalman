import json
import os
from fastapi import HTTPException
import openpyxl

DB_FILE = "tasks_db.json"
OUTPUT_DIR = "outputs"

def load_db():
    if not os.path.exists(DB_FILE):
        return {}
    with open(DB_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def apply_bulk_correction(task_id: str, corrections: list):
    db = load_db()
    task = db.get(task_id)
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
        
    if task["status"] != "completed":
        raise HTTPException(status_code=400, detail="Cannot correct a task that is not completed")

    files = task.get("files", [])
    
    doc_mapping = {
        "bom":     "Bill of Materials",
        "pfd":     "Engineering PFD",
        "tooling": "Tooling Master List",
        "fitment": "Fitment Checksheet",
        "pfmea_assembly": "Process FMEA",
        "pfmea_sheetmetal": "Process FMEA",
        "pfmea_bop": "Process FMEA",
        "control_plan": "Control Plan",
        "fixture_plan": "Fixture PM Master Plan",
    }

    # Group corrections by document
    grouped_corrections = {}
    for c in corrections:
        doc = c["document"]
        if doc not in grouped_corrections:
            grouped_corrections[doc] = []
        grouped_corrections[doc].append(c)

    modified_file_urls = []

    for doc_type, edits in grouped_corrections.items():
        expected_name = doc_mapping.get(doc_type)
        if not expected_name:
            continue
            
        target_file = None
        for f in files:
            if expected_name in f["name"]:
                target_file = f["url"]
                break
                
        if not target_file:
            continue
            
        file_path = target_file.lstrip("/")
        
        if not os.path.exists(file_path):
            continue

        try:
            wb = openpyxl.load_workbook(file_path)
            
            for edit in edits:
                cell_no = edit["cell_no"]
                column = edit["column"]
                replacement_content = edit["replacement_content"]
                
                target_sheet = None
                if doc_type in ["bom", "tooling", "fitment", "fixture_plan"]:
                    # Single-sheet documents: write directly to the active sheet
                    target_sheet = wb.active
                elif doc_type == "pfd":
                    if column.startswith("pfd_1"):
                        sheet_name = "Sub Assmbly Index"
                    elif column.startswith("pfd_2"):
                        sheet_name = "SUB_ASSY"
                    elif column.startswith("pfd_3"):
                        sheet_name = "Q-SHEETMETAL_PARTS"
                    elif column.startswith("pfd_4"):
                        sheet_name = "T - BOP & Hardwares"
                    else:
                        continue

                    if sheet_name in wb.sheetnames:
                        target_sheet = wb[sheet_name]
                elif doc_type == "control_plan":
                    if column.startswith("cp_1"):
                        sheet_name = "Sub Assmbly Index"
                    elif column.startswith("cp_2"):
                        sheet_name = "ASSY_SUB_ASSY"
                    elif column.startswith("cp_3"):
                        sheet_name = "Q-Sheetmetal Parts"
                    elif column.startswith("cp_4"):
                        sheet_name = "T - BOP & Hardwares"
                    else:
                        continue

                    if sheet_name in wb.sheetnames:
                        target_sheet = wb[sheet_name]
                elif doc_type.startswith("pfmea"):
                    if doc_type == "pfmea_assembly":
                        sheet_name = "Assembly PFMEA"
                    elif doc_type == "pfmea_sheetmetal":
                        sheet_name = "Sheet Metal Parts PFMEA"
                    elif doc_type == "pfmea_bop":
                        sheet_name = "BOP Parts PFMEA"
                    else:
                        continue

                    if sheet_name in wb.sheetnames:
                        target_sheet = wb[sheet_name]
                
                if target_sheet:
                    target_sheet[cell_no].value = replacement_content
            
            wb.save(file_path)
            modified_file_urls.append({"document": doc_type, "url": target_file})
            
        except Exception as e:
            print(f"Error applying correction for {doc_type}: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to modify Excel document: {str(e)}")

    return {
        "status": "success", 
        "message": f"Successfully updated {sum(len(edits) for edits in grouped_corrections.values())} cell(s).",
        "modified_files": modified_file_urls
    }
