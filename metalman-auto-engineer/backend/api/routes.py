from fastapi import APIRouter, UploadFile, File, BackgroundTasks, Form, HTTPException
import os
import shutil
import uuid
import json
import re
import pandas as pd
from pydantic import BaseModel
from typing import List
from services.pfd_renderer import render_sub_assy_index, render_sub_assy_sheet, render_sheetmetal_sheet, render_bop_sheet
from services.cad_service import export_step_to_stl
from services.bom_generator import generate_bom_excel
from services.tooling_generator import generate_tooling_list
from services.fitment_generator import generate_fitment_check_sheet
from services.pfmea_generator import generate_pfmea_excel
from services.control_plan_generator import generate_control_plan_excel
from services.fixture_plan_generator import generate_fixture_plan
from services.correction_service import apply_bulk_correction
from xlsx2html import xlsx2html
from fastapi.responses import HTMLResponse, FileResponse
import io
try:
    import pythoncom
    import win32com.client
    HAS_WIN32 = True
except ImportError:
    pythoncom = None
    HAS_WIN32 = False

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import base64
from io import BytesIO
router = APIRouter()

@router.get("/demo-folders")
async def get_demo_folders():
    # Path to the internal stitch_suite folder
    demo_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "stitch_suite"))
    folders = []
    if os.path.exists(demo_path):
        for item in os.listdir(demo_path):
            if os.path.isdir(os.path.join(demo_path, item)) and not item.startswith("."):
                # Filter out landing and login pages as requested
                if "landing" in item or "login" in item:
                    continue
                # Check if content_only.html or code.html exists
                if os.path.exists(os.path.join(demo_path, item, "content_only.html")) or \
                   os.path.exists(os.path.join(demo_path, item, "code.html")):
                    folders.append(item)
    return sorted(folders)

class CorrectionItem(BaseModel):
    document: str
    column: str
    cell_no: str
    replacement_content: str

class BulkCorrectionRequest(BaseModel):
    task_id: str
    corrections: List[CorrectionItem]


DATA_DIR = "/data" if os.path.exists("/data") else os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
UPLOAD_DIR = os.path.abspath(os.path.join(DATA_DIR, "uploads"))
OUTPUT_DIR = os.path.abspath(os.path.join(DATA_DIR, "outputs"))
TASKS_FILE = os.path.join(DATA_DIR, "tasks_db.json")
WELDING_EXCEL_FILENAME = os.path.join(OUTPUT_DIR, "welding_crops.xlsx")
PFD_TEMPLATE_PATH     = os.path.join("assets", "pfd template.xlsx")
BOM_TEMPLATE_PATH     = os.path.join("assets", "BLANK_BOM_TEMPLATE.xlsx")
TOOLING_TEMPLATE_PATH = os.path.join("assets", "tooling list temp.xlsx")
FITMENT_TEMPLATE_PATH = os.path.join("assets", "fitment_checksheet_template.xlsx")
PFMEA_TEMPLATE_PATH   = os.path.join("assets", "PFMEA_TEMPLATE.xlsx")
PFMEA_DICT_DIR             = os.path.join("assets", "pfmea_dicts")
CONTROL_PLAN_TEMPLATE_PATH = os.path.join("assets", "CONTROL_PLAN_TEMPLATE.xlsx")
CONTROL_PLAN_DICT_DIR           = os.path.join("assets", "control_plan_dicts")
FIXTURE_PLAN_TEMPLATE_PATH     = os.path.join("assets", "fixture_plan_template.xlsx")

# Task tracking
TASKS = {} # task_id -> {status, progress, stl_url, files: [], message, error}

def load_tasks():
    global TASKS
    if os.path.exists(TASKS_FILE):
        try:
            with open(TASKS_FILE, "r") as f:
                TASKS = json.load(f)
        except:
            TASKS = {}

def save_tasks():
    try:
        with open(TASKS_FILE, "w") as f:
            json.dump(TASKS, f, indent=2)
    except:
        pass

# Ensure directories exist
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Welding Cropper State
current_welding_row = 1

def init_welding_excel():
    global current_welding_row
    try:
        if os.path.exists(WELDING_EXCEL_FILENAME):
            wb = load_workbook(WELDING_EXCEL_FILENAME)
            ws = wb.active
            current_welding_row = ws.max_row + 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "welding sheet"
            ws.cell(row=1, column=1, value="ID")
            ws.cell(row=1, column=2, value="Cropped Image")
            wb.save(WELDING_EXCEL_FILENAME)
            current_welding_row = 2
    except Exception as e:
        print(f"[WELDING] Init error: {e}")
        current_welding_row = 2

init_welding_excel()

load_tasks()

# Ensure directories exist
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

def process_engineering_task(task_id: str, cad_path: str, feasibility_path: str, drawings_dir: str, requested_outputs: dict):
    """Full pipeline: 3D -> BOM -> PFD -> Tooling."""
    try:
        print(f"\n[TASK {task_id}] Starting engineering pipeline...")
        TASKS[task_id]["status"] = "processing"
        TASKS[task_id]["files"] = []
        save_tasks()
        
        # 0. Extract Part Number from Feasibility for proper naming
        print(f"[TASK {task_id}] Extracting Part Number for naming convention...")
        part_no = "UNKNOWN"
        try:
            # 🛡️ PRIORITY 1: Filename (usually contains the Master Part Number)
            match = re.search(r'\d{8,9}', os.path.basename(feasibility_path))
            if match:
                part_no = match.group(0)
            else:
                # PRIORITY 2: Inside Excel (might be a sub-part, so secondary)
                df_temp = pd.read_excel(feasibility_path, sheet_name=0, skiprows=3)
                col_match = [c for c in df_temp.columns if 'PART NO' in str(c).upper()]
                if col_match:
                    val = df_temp[col_match[0]].dropna().iloc[0]
                    part_no = str(val).replace('.0', '').strip()
                else:
                    # PRIORITY 3: Any digits in filename
                    match = re.search(r'\d+', os.path.basename(feasibility_path))
                    if match: part_no = match.group(0)
        except Exception as e:
            print(f"   [WARN] Could not extract part number: {e}")
            
        print(f"[TASK {task_id}] Target Part Number: {part_no}")
        
        # 1. Export 3D STL
        TASKS[task_id]["progress"] = 5
        TASKS[task_id]["message"] = "Initializing CAD geometry..."
        print(f"[TASK {task_id}] Exporting STEP to STL: {cad_path}")
        stl_filename = f"model_{part_no}.stl"
        stl_path = os.path.join(OUTPUT_DIR, stl_filename)
        if export_step_to_stl(cad_path, stl_path):
            TASKS[task_id]["stl_url"] = f"/outputs/{stl_filename}"
            TASKS[task_id]["progress"] = 15
            save_tasks()
            print(f"[TASK {task_id}] STL export successful.")
        
        # 2. BOM Generation & Image Extraction
        # Note: Fitment Checksheet also relies on images extracted during this phase.
        if requested_outputs.get("bom") or requested_outputs.get("fitment") or requested_outputs.get("control_plan"):
            print(f"[TASK {task_id}] Starting BOM/Image generation...")
            TASKS[task_id]["message"] = "Generating Bill of Materials & Extracting Images..."
            TASKS[task_id]["progress"] = 20
            
            bom_filename = f"BOM_Result_{part_no}.xlsx"
            bom_path = os.path.join(OUTPUT_DIR, bom_filename)
            
            tmp_img_dir = os.path.join(UPLOAD_DIR, f"tmp_imgs_{task_id}")
            os.makedirs(tmp_img_dir, exist_ok=True)
            
            generate_bom_excel(
                feasibility_path, 
                BOM_TEMPLATE_PATH, 
                drawings_dir, 
                tmp_img_dir, 
                bom_path
            )
            
            if requested_outputs.get("bom"):
                TASKS[task_id]["files"].append({"name": "Bill of Materials", "url": f"/outputs/{bom_filename}", "type": "xlsx"})
            
            TASKS[task_id]["progress"] = 45
            save_tasks()
            print(f"[TASK {task_id}] BOM/Image generation complete.")

        # 3. PFD Generation
        if requested_outputs.get("pfd"):
            print(f"[TASK {task_id}] Starting PFD generation pipeline...")
            TASKS[task_id]["message"] = "Generating Progressive Flow Diagrams..."
            excel_filename = f"PFD_Result_{part_no}.xlsx"
            excel_path = os.path.join(OUTPUT_DIR, excel_filename)
            
            shutil.copy(PFD_TEMPLATE_PATH, excel_path)
            
            print(f"[TASK {task_id}] Rendering PFD Sheet 1 (Sub-Assy Index)...")
            render_sub_assy_index(feasibility_path, cad_path, excel_path, excel_path, OUTPUT_DIR)
            TASKS[task_id]["progress"] = 60
            
            print(f"[TASK {task_id}] Rendering PFD Sheet 2 (Sub-Assy Flow)...")
            render_sub_assy_sheet(feasibility_path, excel_path, excel_path)
            TASKS[task_id]["progress"] = 70
            
            print(f"[TASK {task_id}] Rendering PFD Sheet 3 (Sheetmetal Flow)...")
            render_sheetmetal_sheet(feasibility_path, excel_path, excel_path)
            TASKS[task_id]["progress"] = 80
            
            print(f"[TASK {task_id}] Rendering PFD Sheet 4 (BOP List)...")
            render_bop_sheet(feasibility_path, excel_path, excel_path)
            
            TASKS[task_id]["files"].append({"name": "Engineering PFD (Excel)", "url": f"/outputs/{excel_filename}", "type": "xlsx"})
            TASKS[task_id]["progress"] = 90
            print(f"[TASK {task_id}] PFD generation complete.")

        # 4. Tooling Analysis (Fixture Master List)
        if requested_outputs.get("tooling"):
            print(f"[TASK {task_id}] Starting Tooling List generation...")
            TASKS[task_id]["message"] = "Generating Tooling Master List..."
            TASKS[task_id]["progress"] = 92
            save_tasks()
            
            tool_filename = f"Tooling_List_{part_no}.xlsx"
            tool_path = os.path.join(OUTPUT_DIR, tool_filename)
            tool_img_dir = os.path.join(UPLOAD_DIR, f"tool_imgs_{task_id}")
            
            if generate_tooling_list(feasibility_path, TOOLING_TEMPLATE_PATH, tool_path, tool_img_dir):
                TASKS[task_id]["files"].append({"name": "Tooling Master List", "url": f"/outputs/{tool_filename}", "type": "xlsx"})
            
            TASKS[task_id]["progress"] = 95
            save_tasks()
            print(f"[TASK {task_id}] Tooling List generation complete.")

        # 5. Fitment Checksheet
        if requested_outputs.get("fitment"):
            print(f"[TASK {task_id}] Starting Fitment Checksheet generation...")
            TASKS[task_id]["message"] = "Generating Fitment Checksheet..."
            TASKS[task_id]["progress"] = 97
            save_tasks()

            fitment_filename = f"Fitment_Checksheet_{part_no}.xlsx"
            fitment_path = os.path.join(OUTPUT_DIR, fitment_filename)

            # Use the shared image directory (ensured to be populated by Step 2)
            tmp_img_dir = os.path.join(UPLOAD_DIR, f"tmp_imgs_{task_id}")
            
            if generate_fitment_check_sheet(feasibility_path, FITMENT_TEMPLATE_PATH, fitment_path, tmp_img_dir):
                TASKS[task_id]["files"].append({"name": "Fitment Checksheet", "url": f"/outputs/{fitment_filename}", "type": "xlsx"})

            save_tasks()
            print(f"[TASK {task_id}] Fitment Checksheet generation complete.")

        # 6. PFMEA Generation
        if requested_outputs.get("pfmea"):
            print(f"[TASK {task_id}] Starting PFMEA generation...")
            TASKS[task_id]["message"] = "Generating PFMEA Document..."
            TASKS[task_id]["progress"] = 98
            save_tasks()

            pfmea_filename = f"PFMEA_Result_{part_no}.xlsx"
            pfmea_path = os.path.join(OUTPUT_DIR, pfmea_filename)

            if generate_pfmea_excel(feasibility_path, PFMEA_TEMPLATE_PATH, pfmea_path, PFMEA_DICT_DIR):
                TASKS[task_id]["files"].append({"name": "Process FMEA", "url": f"/outputs/{pfmea_filename}", "type": "xlsx"})
            
            save_tasks()
            print(f"[TASK {task_id}] PFMEA generation complete.")

        # 7. Control Plan Generation
        if requested_outputs.get("control_plan"):
            print(f"[TASK {task_id}] Starting Control Plan generation...")
            TASKS[task_id]["message"] = "Generating Control Plan..."
            TASKS[task_id]["progress"] = 99
            save_tasks()

            cp_filename = f"Control_Plan_{part_no}.xlsx"
            cp_path     = os.path.join(OUTPUT_DIR, cp_filename)

            # Resolve paths needed by the generator
            tmp_img_dir  = os.path.join(UPLOAD_DIR, f"tmp_imgs_{task_id}")
            pfd_filename = f"PFD_Result_{part_no}.xlsx"
            pfd_out_path = os.path.join(OUTPUT_DIR, pfd_filename)

            # Build sub-assy index data from feasibility (reuse same extraction logic)
            try:
                from services.pfd_renderer import extract_feasibility_data
                _, assy_data = extract_feasibility_data(feasibility_path)
            except Exception as e:
                print(f"[CP] Failed to extract sub assy data: {e}")
                assy_data = []

            generated = generate_control_plan_excel(
                feasibility_path=feasibility_path,
                template_path=CONTROL_PLAN_TEMPLATE_PATH,
                output_path=cp_path,
                dict_dir=CONTROL_PLAN_DICT_DIR,
                pfd_output_path=pfd_out_path,
                img_dir=tmp_img_dir,
                assy_data=assy_data,
                pfd_img_dir=OUTPUT_DIR,
            )
            if generated:
                TASKS[task_id]["files"].append({"name": "Control Plan", "url": f"/outputs/{cp_filename}", "type": "xlsx"})

            save_tasks()
            print(f"[TASK {task_id}] Control Plan generation complete.")

        # 8. Fixture PM Master Plan
        if requested_outputs.get("fixture_plan"):
            print(f"[TASK {task_id}] Starting Fixture Plan generation...")
            TASKS[task_id]["message"] = "Generating Fixture PM Master Plan..."
            TASKS[task_id]["progress"] = 99.5
            save_tasks()

            fp_filename = f"Fixture_Plan_{part_no}.xlsx"
            fp_path     = os.path.join(OUTPUT_DIR, fp_filename)

            if generate_fixture_plan(feasibility_path, FIXTURE_PLAN_TEMPLATE_PATH, fp_path):
                TASKS[task_id]["files"].append({"name": "Fixture PM Master Plan", "url": f"/outputs/{fp_filename}", "type": "xlsx"})

            save_tasks()
            print(f"[TASK {task_id}] Fixture Plan generation complete.")

        TASKS[task_id]["progress"] = 100
        TASKS[task_id]["status"] = "completed"
        TASKS[task_id]["message"] = "Generation complete"
        save_tasks()
        print(f"[TASK {task_id}] ALL TASKS COMPLETED SUCCESSFULLY.\n")
        
    except Exception as e:
        print(f"[TASK {task_id}] FATAL ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        TASKS[task_id]["status"] = "error"
        TASKS[task_id]["error"] = str(e)
        TASKS[task_id]["message"] = f"Error: {str(e)}"

@router.post("/analyze")
async def start_analysis(
    background_tasks: BackgroundTasks,
    cad_file: UploadFile = File(...),
    feasibility_file: UploadFile = File(...),
    drawings: list[UploadFile] = File(default=[]),
    outputs: str = Form("{}")
):
    task_id = str(uuid.uuid4())[:8]
    requested_outputs = json.loads(outputs)
    
    # Initialize task tracking
    TASKS[task_id] = {
        "status": "queued",
        "progress": 0,
        "message": "Task queued...",
        "stl_url": None,
        "files": [],
        "error": None
    }
    save_tasks()
    
    # Save files
    cad_path = os.path.join(UPLOAD_DIR, f"{task_id}_{cad_file.filename}")
    feas_path = os.path.join(UPLOAD_DIR, f"{task_id}_{feasibility_file.filename}")
    
    with open(cad_path, "wb") as f:
        shutil.copyfileobj(cad_file.file, f)
    with open(feas_path, "wb") as f:
        shutil.copyfileobj(feasibility_file.file, f)
        
    # Save drawings if any
    drawings_dir = os.path.join(UPLOAD_DIR, f"{task_id}_drawings")
    os.makedirs(drawings_dir, exist_ok=True)
    for drawing in drawings:
        d_path = os.path.join(drawings_dir, drawing.filename)
        with open(d_path, "wb") as f:
            shutil.copyfileobj(drawing.file, f)
            
    # Start background task
    background_tasks.add_task(process_engineering_task, task_id, cad_path, feas_path, drawings_dir, requested_outputs)
    
    return {
        "message": "Analysis started",
        "task_id": task_id
    }

@router.get("/status/{task_id}")
async def get_analysis_status(task_id: str):
    if task_id in TASKS:
        return TASKS[task_id]
    return {"status": "not_found"}

@router.post("/correct")
async def correct_document(req: BulkCorrectionRequest):
    return apply_bulk_correction(
        task_id=req.task_id,
        corrections=[c.dict() for c in req.corrections]
    )

@router.get("/preview/{filename}")
async def preview_excel(filename: str):
    """
    High-fidelity 'Sure Shot' preview solution.
    """
    excel_path = os.path.join(OUTPUT_DIR, filename)
    pdf_filename = filename.replace(".xlsx", ".pdf")
    pdf_path = os.path.join(OUTPUT_DIR, pdf_filename)
    
    # 1. IMMEDIATE RETURN: If PDF already exists, serve it instantly
    if os.path.exists(pdf_path):
        return FileResponse(pdf_path, media_type="application/pdf")

    if not os.path.exists(excel_path):
        raise HTTPException(status_code=404, detail=f"File not found: {filename}")
    
    # 2. GENERATE PDF: Try native Excel conversion
    if HAS_WIN32:
        try:
            pythoncom.CoInitialize()
            # Use DispatchEx for a new instance and better isolation
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            try:
                wb = excel.Workbooks.Open(excel_path)
                # Ensure all sheets are processed for full visibility
                for sheet in wb.Sheets:
                    try:
                        # COMPUTE TRUE DATA RANGE (Ignoring rows with only formatting/borders)
                        # xlByRows=1, xlPrevious=2, xlValues=18 (optional to be strict)
                        try:
                            res_row = sheet.Cells.Find(What="*", SearchOrder=1, SearchDirection=2)
                            res_col = sheet.Cells.Find(What="*", SearchOrder=2, SearchDirection=2) # xlByColumns=2
                            
                            if res_row and res_col:
                                last_row = res_row.Row
                                last_col = res_col.Column
                                
                                # Safety: don't clip headers if data is small, usually starts at 1,1
                                start_cell = sheet.Cells(1, 1).Address
                                end_cell   = sheet.Cells(last_row, last_col).Address
                                sheet.PageSetup.PrintArea = f"{start_cell}:{end_cell}"
                            else:
                                sheet.PageSetup.PrintArea = "$A$1:$F$20" # Fallback to a small header area
                        except Exception as find_err:
                            print(f"[PREVIEW] Find error: {find_err}")
                            sheet.PageSetup.PrintArea = ""

                        # Ensure everything fits on one page width but can span multiple pages down
                        sheet.PageSetup.Zoom = False
                        sheet.PageSetup.FitToPagesWide = 1
                        sheet.PageSetup.FitToPagesTall = False
                    except Exception as sheet_err:
                        print(f"[PREVIEW] Sheet setup error: {sheet_err}")
                
                # Export the entire workbook to PDF
                # 0 = xlTypePDF, 1 = xlQualityStandard, IncludeDocProperties=True, IgnorePrintAreas=False (Crucial to respect the PrintArea we set)
                wb.ExportAsFixedFormat(0, pdf_path, Quality=0, IncludeDocProperties=True, IgnorePrintAreas=False) 
                wb.Close(False)
            finally:
                excel.Quit()
            
            if os.path.exists(pdf_path):
                return FileResponse(pdf_path, media_type="application/pdf")
        except Exception as e:
            print(f"[PREVIEW ERROR] Native conversion failed: {e}")
        finally:
            try: pythoncom.CoUninitialize()
            except: pass

    # 3. FALLBACK: Data View (if PDF fails)
    try:
        # Use pandas for a clean data-only fallback instead of xlsx2html which is buggy with images
        df = pd.read_excel(excel_path, sheet_name=0)
        html_content = df.to_html(classes="table table-striped", index=False)
        
        styled_html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: 'Segoe UI', sans-serif; padding: 20px; background: #f0f2f5; }}
                .container {{ background: white; padding: 30px; border-radius: 8px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); margin: auto; }}
                table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
                td, th {{ border: 1px solid #dfe3e8; padding: 10px; text-align: left; }}
                th {{ background-color: #f4f6f8; font-weight: 600; color: #454f5b; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h3 style="margin-top:0">Data Preview (Low Fidelity)</h3>
                {html_content}
            </div>
        </body>
        </html>
        """
        return HTMLResponse(content=styled_html)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Preview failed: {str(e)}")

@router.post("/save-crop")
async def save_crop(image_data: str = Form(...)):
    global current_welding_row
    try:
        # Decode base64 image
        header, encoded = image_data.split(",", 1)
        data = base64.b64decode(encoded)
        img_buffer = BytesIO(data)
        image = Image.open(img_buffer)
        
        # Save image to outputs
        crop_id = str(uuid.uuid4())[:8]
        crop_filename = f"crop_{crop_id}.png"
        crop_path = os.path.join(OUTPUT_DIR, crop_filename)
        image.save(crop_path)
        
        # Load existing Excel
        wb = load_workbook(WELDING_EXCEL_FILENAME)
        ws = wb.active
        
        # Add ID
        ws.cell(row=current_welding_row, column=1, value=current_welding_row - 1)
        
        # Insert image in Column B
        img = OpenpyxlImage(crop_path)
        img.width = 250
        img.height = 180
        
        # Adjust row height to fit image
        ws.row_dimensions[current_welding_row].height = 150
        ws.column_dimensions['B'].width = 40
        
        ws.add_image(img, f"B{current_welding_row}")
        
        wb.save(WELDING_EXCEL_FILENAME)
        current_welding_row += 1
        
        return {"status": "success", "message": f"Image saved in row {current_welding_row-1}", "row": current_welding_row-1}
    except Exception as e:
        print(f"[WELDING] Save error: {e}")
        return {"status": "error", "message": str(e)}

@router.get("/download-excel")
async def download_welding_excel():
    if os.path.exists(WELDING_EXCEL_FILENAME):
        return FileResponse(WELDING_EXCEL_FILENAME, filename="welding_report.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    raise HTTPException(status_code=404, detail="File not found")

@router.post("/reset-excel")
async def reset_welding_excel():
    global current_welding_row
    if os.path.exists(WELDING_EXCEL_FILENAME):
        os.remove(WELDING_EXCEL_FILENAME)
    init_welding_excel()
    return {"status": "success", "message": "Excel has been reset"}
