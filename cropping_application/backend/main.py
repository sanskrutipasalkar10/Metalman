from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import os
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import base64
from PIL import Image
import uuid

app = FastAPI()

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "uploads"
OUT_DIR = "output"
if not os.path.exists(UPLOAD_DIR): os.makedirs(UPLOAD_DIR)
if not os.path.exists(OUT_DIR): os.makedirs(OUT_DIR)

# Global or session-based state for the current excel
# For simplicity, we'll use a single global file for this demo
EXCEL_FILENAME = os.path.join(OUT_DIR, "welding_crops.xlsx")
current_row = 1

def init_excel():
    global current_row
    wb = Workbook()
    ws = wb.active
    ws.title = "welding sheet"
    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=2, value="Cropped Image")
    wb.save(EXCEL_FILENAME)
    current_row = 2

# Initialize on start
init_excel()

@app.get("/")
def read_root():
    return {"message": "PDF Cropping Backend Running"}

@app.post("/upload-pdf")
async def upload_pdf(file: UploadFile = File(...)):
    file_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return {"filename": file.filename, "path": file_path}

@app.post("/save-crop")
async def save_crop(
    image_data: str = Form(...)
):
    global current_row
    try:
        # Decode base64 image
        header, encoded = image_data.split(",", 1)
        data = base64.b64decode(encoded)
        img_buffer = BytesIO(data)
        image = Image.open(img_buffer)
        
        # Save temporary image for openpyxl
        temp_id = str(uuid.uuid4())
        temp_img_path = os.path.join(OUT_DIR, f"crop_{temp_id}.png")
        image.save(temp_img_path)
        
        # Load existing Excel
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb["welding sheet"]
        
        # Add ID
        ws.cell(row=current_row, column=1, value=current_row - 1)
        
        # Insert image in Column B
        img = OpenpyxlImage(temp_img_path)
        # Resize display size in Excel (higher resolution source stays intact)
        img.width = 250
        img.height = 180
        
        # Adjust row height to fit image
        ws.row_dimensions[current_row].height = 150
        ws.column_dimensions['B'].width = 40
        
        ws.add_image(img, f"B{current_row}")
        
        wb.save(EXCEL_FILENAME)
        current_row += 1
        
        return {"status": "success", "message": f"Image saved in row {current_row-1}", "row": current_row-1}
    except Exception as e:
        print(f"Error: {e}")
        return {"status": "error", "message": str(e)}

@app.get("/download-excel")
async def download_excel():
    if os.path.exists(EXCEL_FILENAME):
        return FileResponse(EXCEL_FILENAME, filename="welding_report.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    raise HTTPException(status_code=404, detail="File not found")

@app.post("/reset-excel")
async def reset_excel():
    init_excel()
    return {"status": "success", "message": "Excel has been reset"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
